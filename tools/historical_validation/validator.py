from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from loan_interest_accrual.domain import (
    DayCountBasis,
    Loan,
    NaturalMonth,
    calculate_loan,
)


ALLOWED_REASON_CODES = frozenset(("match", "start_date_excluded"))


@dataclass(frozen=True, slots=True)
class _ManifestPath:
    def read_text(self, encoding: str = "utf-8") -> str:
        return json.dumps(MANIFEST, ensure_ascii=False, indent=2, sort_keys=True)


MANIFEST: dict[str, Any] = {
    "version": 1,
    "cases": [
        {
            "case_id": "hist-2024-01-row2-match",
            "calculation_month": "2024-01",
            "source_evidence": [
                {"sheet": "24全年", "cell": "A2", "meaning": "source row id"},
                {"sheet": "24全年", "cell": "B2", "meaning": "borrower"},
                {"sheet": "24全年", "cell": "C2", "meaning": "borrow date"},
                {"sheet": "24全年", "cell": "D2", "meaning": "principal"},
                {"sheet": "24全年", "cell": "E1", "meaning": "rate and basis label"},
                {"sheet": "24全年", "cell": "E2", "meaning": "daily interest formula"},
                {"sheet": "24全年", "cell": "F2", "meaning": "interest days"},
                {"sheet": "24全年", "cell": "J2", "meaning": "day count basis"},
            ],
            "historical_value_cell": {"sheet": "24全年", "cell": "G2"},
            "canonical_input": {
                "loan": {
                    "loan_id": "hist-2024-01-row2-match",
                    "company_name": "上海翰凌医疗器械有限公司",
                    "contract_number": "historical-24-full-year-row-2",
                    "bank_name": "historical workbook",
                    "opening_principal": "5000000",
                    "annual_rate": "0.025",
                    "day_count_basis": 360,
                    "accrual_start": "2022-11-10",
                    "accrual_end": "2024-01-31",
                    "capitalize_interest": False,
                }
            },
            "expected_reason_code": "match",
        },
        {
            "case_id": "hist-2022-11-row3-rounding-reference",
            "calculation_month": "2022-11",
            "source_evidence": [
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "A3",
                    "meaning": "source row id",
                },
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "B3",
                    "meaning": "borrower",
                },
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "C3",
                    "meaning": "borrow date",
                },
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "D3",
                    "meaning": "principal",
                },
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "E1",
                    "meaning": "rate and basis label",
                },
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "E3",
                    "meaning": "daily interest formula",
                },
                {
                    "sheet": "子公司借款利息（农商）",
                    "cell": "F3",
                    "meaning": "interest days",
                },
            ],
            "historical_value_cell": {
                "sheet": "子公司借款利息（农商）",
                "cell": "G3",
            },
            "canonical_input": {
                "loan": {
                    "loan_id": "hist-2022-11-row3-rounding-reference",
                    "company_name": "上海翰凌医疗器械有限公司",
                    "contract_number": "historical-nongshang-row-3",
                    "bank_name": "historical workbook",
                    "opening_principal": "5000000",
                    "annual_rate": "0.025",
                    "day_count_basis": 360,
                    "accrual_start": "2022-11-10",
                    "accrual_end": "2022-11-30",
                    "capitalize_interest": False,
                }
            },
            "expected_reason_code": "start_date_excluded",
        },
    ],
}

MANIFEST_PATH = _ManifestPath()


class HistoricalValidationError(RuntimeError):
    pass


def _sha256_file(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def _month(value: str) -> NaturalMonth:
    year_text, month_text = value.split("-", 1)
    return NaturalMonth(int(year_text), int(month_text))


def _date(value: str) -> date:
    return date.fromisoformat(value)


def _decimal(value: object) -> Decimal:
    if type(value) is bool or value is None:
        raise HistoricalValidationError(f"expected numeric value, got {value!r}")
    return Decimal(str(value))


def _json_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return value


def _cell_value(workbook, sheet_name: str, cell: str) -> object:
    if sheet_name not in workbook.sheetnames:
        raise HistoricalValidationError(f"missing historical sheet: {sheet_name}")
    return workbook[sheet_name][cell].value


def _evidence_item(data_workbook, formula_workbook, item: dict[str, str]) -> dict[str, object]:
    sheet_name = item["sheet"]
    cell = item["cell"]
    data_value = _cell_value(data_workbook, sheet_name, cell)
    formula_value = _cell_value(formula_workbook, sheet_name, cell)
    evidence: dict[str, object] = {
        "sheet": sheet_name,
        "cell": cell,
        "meaning": item["meaning"],
        "value": _json_value(data_value),
    }
    if formula_value != data_value:
        evidence["formula"] = _json_value(formula_value)
    return evidence


def _loan(values: dict[str, object]) -> Loan:
    return Loan(
        loan_id=str(values["loan_id"]),
        company_name=str(values["company_name"]),
        contract_number=str(values["contract_number"]),
        bank_name=str(values["bank_name"]),
        opening_principal=Decimal(str(values["opening_principal"])),
        annual_rate=Decimal(str(values["annual_rate"])),
        day_count_basis=DayCountBasis(values["day_count_basis"]),
        accrual_start=_date(str(values["accrual_start"])),
        accrual_end=_date(str(values["accrual_end"])),
        capitalize_interest=values["capitalize_interest"] is True,
    )


def _reason_code(delta: Decimal, expected: str) -> str:
    if expected not in ALLOWED_REASON_CODES:
        raise HistoricalValidationError(f"unknown reason code: {expected}")
    if delta == Decimal("0") and expected != "match":
        raise HistoricalValidationError("manifest expected a difference but delta is zero")
    if delta != Decimal("0") and expected == "match":
        raise HistoricalValidationError("manifest expected a match but delta is non-zero")
    return expected


def _case_report(data_workbook, formula_workbook, case: dict[str, Any]) -> dict[str, object]:
    period = _month(case["calculation_month"])
    loan = _loan(case["canonical_input"]["loan"])
    result = calculate_loan(period, loan, ())

    historical_cell = case["historical_value_cell"]
    historical_raw = _cell_value(
        data_workbook,
        historical_cell["sheet"],
        historical_cell["cell"],
    )
    historical_value = _decimal(historical_raw)
    delta = result.accrued_interest - historical_value
    reason_code = _reason_code(delta, case["expected_reason_code"])

    return {
        "case_id": case["case_id"],
        "calculation_month": case["calculation_month"],
        "source_evidence": [
            _evidence_item(data_workbook, formula_workbook, item)
            for item in case["source_evidence"]
        ],
        "canonical_input": case["canonical_input"],
        "canonical_result": {
            "interest_days": result.interest_days,
            "unrounded_interest": format(result.unrounded_interest, "f"),
            "accrued_interest": format(result.accrued_interest, "f"),
        },
        "historical_value": {
            "sheet": historical_cell["sheet"],
            "cell": historical_cell["cell"],
            "value": format(historical_value, "f"),
        },
        "delta": format(delta, "f"),
        "reason_code": reason_code,
    }


def run_validation(source_path: Path, output_dir: Path) -> dict[str, object]:
    source_path = Path(source_path)
    output_dir = Path(output_dir)
    source_hash_before = _sha256_file(source_path)

    data_workbook = load_workbook(source_path, read_only=True, data_only=True)
    formula_workbook = load_workbook(source_path, read_only=True, data_only=False)
    try:
        cases = [
            _case_report(data_workbook, formula_workbook, case)
            for case in MANIFEST["cases"]
        ]
    finally:
        data_workbook.close()
        formula_workbook.close()

    source_hash_after = _sha256_file(source_path)
    if source_hash_after != source_hash_before:
        raise HistoricalValidationError("historical workbook hash changed during validation")

    report = {
        "report_version": 1,
        "source_path": source_path.as_posix(),
        "source_sha256_before": source_hash_before,
        "source_sha256_after": source_hash_after,
        "allowed_reason_codes": sorted(ALLOWED_REASON_CODES),
        "manifest": MANIFEST,
        "cases": cases,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "source-hash-before.txt").write_text(
        f"{source_hash_before}\n",
        encoding="utf-8",
    )
    (output_dir / "source-hash-after.txt").write_text(
        f"{source_hash_after}\n",
        encoding="utf-8",
    )
    (output_dir / "differential-report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Validate selected historical rows.")
    parser.add_argument("source_path", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args(argv)
    run_validation(args.source_path, args.output_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
