from __future__ import annotations

from collections import defaultdict
from dataclasses import replace
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from loan_interest_accrual.application import calculate_submission
from loan_interest_accrual.application.results import ApplicationCheck
from loan_interest_accrual.domain import NaturalMonth
from loan_interest_accrual.workbook.export_schema import (
    CAPITALIZATION_SUMMARY_HEADERS,
    CAPITALIZATION_SUMMARY_SHEET,
    CHECK_HEADERS,
    CHECK_SHEET,
    COMPANY_SUMMARY_HEADERS,
    COMPANY_SUMMARY_SHEET,
    EXPORT_SHEET_NAMES,
    PARAMETER_HEADERS,
    PARAMETER_SHEET,
    RESULT_HEADERS,
    RESULT_SHEET,
    SEGMENT_HEADERS,
    SEGMENT_SHEET,
)
from loan_interest_accrual.workbook.exporter import (
    ExportInvariantError,
    export_calculation_workbook,
)

from tests.fixtures.export.standard_workbooks import (
    loan_row,
    movement_row,
    workbook_bytes,
)


PERIOD = NaturalMonth(2025, 6)
GENERATED_AT = datetime(2026, 7, 10, 12, 30, 0, tzinfo=timezone.utc)


def _decimal(value: object) -> Decimal:
    return value if type(value) is Decimal else Decimal(str(value))


def _money(value: object) -> Decimal:
    return _decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _rows(sheet) -> list[dict[str, object]]:
    headers = [sheet.cell(3, column).value for column in range(1, 9)]
    return [
        dict(zip(headers, values, strict=True))
        for values in sheet.iter_rows(
            min_row=4,
            max_col=8,
            values_only=True,
        )
    ]


def _calculation():
    payload = workbook_bytes(
        loans=[
            loan_row(
                "L-CAP-A",
                company="甲公司",
                contract="HT-A",
                bank="甲银行A",
                principal=1000000,
                rate=0.0365,
                basis=365,
                capitalized="是",
            ),
            loan_row(
                "L-EXP-A",
                company="甲公司",
                contract="HT-B",
                bank="甲银行B",
                principal=2000000,
                rate=0.036,
                basis=360,
                capitalized="否",
            ),
            loan_row(
                "L-CAP-B",
                company="乙公司",
                contract="HT-C",
                bank="乙银行",
                principal=500000,
                rate=0.073,
                basis=365,
                capitalized="是",
            ),
        ],
        movements=[
            movement_row("L-CAP-A", date(2025, 6, 15), "放款", 100000),
            movement_row("L-CAP-A", date(2025, 6, 20), "还本", 50000),
            movement_row("L-EXP-A", date(2025, 6, 30), "放款", 10000),
        ],
    )
    result = calculate_submission("source.xlsx", payload, PERIOD)
    assert result.errors == ()
    assert result.calculation is not None
    return payload, result.calculation


def test_export_workbook_has_exact_sheets_fixed_values_and_package_is_clean() -> None:
    source_bytes, calculation = _calculation()
    output = export_calculation_workbook(
        calculation,
        generated_at=GENERATED_AT,
    )

    assert output != source_bytes
    workbook = load_workbook(BytesIO(output), data_only=False)
    data_only_workbook = load_workbook(BytesIO(output), data_only=True)
    try:
        assert workbook.sheetnames == list(EXPORT_SHEET_NAMES)
        assert data_only_workbook.sheetnames == list(EXPORT_SHEET_NAMES)
        result_sheet = workbook[RESULT_SHEET]
        assert result_sheet["A1"].value == "计提区间  2025-06-01 至 2025-06-30"
        assert [result_sheet.cell(3, column).value for column in range(1, 9)] == list(
            RESULT_HEADERS
        )
        assert [result_sheet.cell(3, column).value for column in range(10, 13)] == [
            "公司名称",
            "本金合计（元）",
            "利息合计（元）",
        ]
        assert {str(merged) for merged in result_sheet.merged_cells.ranges} == {
            "A1:H1",
            "J1:L1",
            "J2:L2",
        }

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    assert cell.data_type != "f"

        assert data_only_workbook[RESULT_SHEET]["A4"].value == 1
        assert data_only_workbook[RESULT_SHEET]["B4"].value == "甲公司"
        assert data_only_workbook[RESULT_SHEET]["C4"].value == "甲银行A"
        assert data_only_workbook[RESULT_SHEET]["J4"].value == "甲公司"
        assert data_only_workbook[RESULT_SHEET]["K4"].value == 3000000
        assert data_only_workbook[RESULT_SHEET]["L4"].value == 9041.67
        assert data_only_workbook[RESULT_SHEET]["J5"].value == "乙公司"
        assert data_only_workbook[RESULT_SHEET]["K5"].value == 500000
        assert data_only_workbook[RESULT_SHEET]["L5"].value == 3041.67
    finally:
        workbook.close()
        data_only_workbook.close()

    with ZipFile(BytesIO(output)) as package:
        names = package.namelist()
        assert not any("vbaProject" in name for name in names)
        assert not any("externalLinks" in name for name in names)
        assert not any("connections" in name for name in names)
        for name in names:
            if not (name.endswith(".xml") or name.endswith(".rels")):
                continue
            content = package.read(name).decode("utf-8", errors="ignore")
            assert "<f>" not in content
            assert "TargetMode=\"External\"" not in content
            assert "source.xlsx" not in content
            assert "贷款主表" not in content


def test_export_rows_summaries_checks_and_parameters_reconcile() -> None:
    _, calculation = _calculation()
    output = export_calculation_workbook(
        calculation,
        generated_at=GENERATED_AT,
    )

    workbook = load_workbook(BytesIO(output), data_only=True)
    try:
        result_rows = _rows(workbook[RESULT_SHEET])
    finally:
        workbook.close()

    assert list(result_rows[0]) == list(RESULT_HEADERS)
    assert len(result_rows) == 4
    assert result_rows == [
        {
            "序号": 1,
            "公司名称": "甲公司",
            "贷款银行": "甲银行A",
            "期初本金（元）": 1000000,
            "年利率": 0.0365,
            "借款时间": datetime(2025, 1, 1),
            "区间应提利息（元）": 3041.67,
            "区间应提利息天数": 30,
        },
        {
            "序号": 2,
            "公司名称": "甲公司",
            "贷款银行": "甲银行B",
            "期初本金（元）": 2000000,
            "年利率": 0.036,
            "借款时间": datetime(2025, 1, 1),
            "区间应提利息（元）": 6000,
            "区间应提利息天数": 30,
        },
        {
            "序号": 3,
            "公司名称": "乙公司",
            "贷款银行": "乙银行",
            "期初本金（元）": 500000,
            "年利率": 0.073,
            "借款时间": datetime(2025, 1, 1),
            "区间应提利息（元）": 3041.67,
            "区间应提利息天数": 30,
        },
        {
            "序号": "合计",
            "公司名称": None,
            "贷款银行": None,
            "期初本金（元）": 3500000,
            "年利率": None,
            "借款时间": None,
            "区间应提利息（元）": 12083.34,
            "区间应提利息天数": None,
        },
    ]


def test_export_is_blocked_when_any_check_has_failed() -> None:
    _, calculation = _calculation()
    failed_calculation = replace(
        calculation,
        checks=calculation.checks
        + (
            ApplicationCheck(
                name="forced_failure",
                display_name="强制失败",
                passed=False,
                expected=Decimal("0"),
                actual=Decimal("1"),
            ),
        ),
    )

    with pytest.raises(ExportInvariantError):
        export_calculation_workbook(failed_calculation, generated_at=GENERATED_AT)
