from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from loan_interest_accrual.application import (
    calculate_amortization_submission,
    export_amortization_submission,
)
from loan_interest_accrual.domain import NaturalMonth
from loan_interest_accrual.workbook import (
    AMORTIZATION_HEADERS,
    AMORTIZATION_SHEET,
    generate_amortization_template,
)


def _input_workbook() -> bytes:
    workbook = load_workbook(BytesIO(generate_amortization_template()))
    sheet = workbook[AMORTIZATION_SHEET]
    sheet.append(
        [
            "软件",
            "已摊满软件",
            "管理费用",
            1000,
            0,
            date(2026, 1, 31),
            date(2026, 1, 31),
            6,
        ]
    )
    sheet.append(
        [
            "软件",
            "未摊满软件",
            "管理费用",
            1200,
            0,
            date(2026, 4, 30),
            date(2026, 4, 30),
            12,
        ]
    )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_amortization_template_matches_required_eight_columns() -> None:
    workbook = load_workbook(BytesIO(generate_amortization_template()))
    try:
        assert workbook.sheetnames == [AMORTIZATION_SHEET]
        sheet = workbook[AMORTIZATION_SHEET]
        assert tuple(cell.value for cell in sheet[1]) == AMORTIZATION_HEADERS
        assert sheet.freeze_panes == "A2"
        assert all(cell.fill.fgColor.rgb.endswith("FFF200") for cell in sheet[1])
        assert sheet["D2"].number_format.startswith("#,##0.00")
        assert sheet["F2"].number_format == "yyyy-mm-dd"
    finally:
        workbook.close()


def test_calculation_applies_completion_cap_and_inclusive_months() -> None:
    result = calculate_amortization_submission(
        "input.xlsx", _input_workbook(), NaturalMonth(2026, 6)
    )
    assert result.errors == ()
    assert result.calculation is not None
    completed, active = result.calculation.rows

    assert completed.sequence == 1
    assert completed.monthly_amortization.as_tuple().exponent == -2
    assert completed.cumulative_months == 6
    assert completed.cumulative_amortization == completed.original_value
    assert completed.current_required_amortization is None
    assert completed.current_actual_amortization is None
    assert completed.difference is None
    assert completed.ending_net_value is None

    assert active.sequence == 2
    assert active.cumulative_months == 3
    assert active.monthly_amortization == 100
    assert active.cumulative_amortization == 300
    assert active.current_required_amortization == 100
    assert active.current_actual_amortization == 100
    assert active.difference is None
    assert active.ending_net_value == 900
    assert all(check.passed for check in result.calculation.checks)


def test_blank_residual_is_optional_and_defaults_to_zero() -> None:
    workbook = load_workbook(BytesIO(generate_amortization_template()))
    sheet = workbook[AMORTIZATION_SHEET]
    sheet.append(
        [
            "软件",
            "无残值软件",
            "管理费用",
            1200,
            None,
            date(2026, 4, 30),
            date(2026, 4, 30),
            12,
        ]
    )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    result = calculate_amortization_submission(
        "blank-residual.xlsx", stream.getvalue(), NaturalMonth(2026, 6)
    )
    assert result.errors == ()
    assert result.calculation is not None
    row = result.calculation.rows[0]
    assert row.residual_value == Decimal("0")
    assert row.monthly_amortization == Decimal("100.00")


def test_export_contains_formulas_and_requested_total_row() -> None:
    result = export_amortization_submission(
        "input.xlsx", _input_workbook(), NaturalMonth(2026, 6)
    )
    assert result.errors == ()
    assert result.output is not None
    workbook = load_workbook(BytesIO(result.output.workbook_bytes), data_only=False)
    try:
        assert workbook.sheetnames == ["摊销明细", "分类明细"]
        assert workbook.active.title == "摊销明细"
        classified = workbook["分类明细"]
        assert [classified.cell(3, column).value for column in range(1, 4)] == [
            "费用",
            "一级分类",
            "当期应摊销金额",
        ]
        sheet = workbook["摊销明细"]
        assert sheet["A1"].value == "无形资产、长期待摊费用摊销明细  2026年6月"
        assert sheet["A2"].value == "计提月份"
        assert sheet["A3"].value == datetime(2026, 6, 1)
        assert sheet["A3"].number_format == 'yy.m"月"'
        assert sheet["B3"].value == 1
        assert sheet["B4"].value == 2
        assert sheet["K3"].value == "=ROUND((F3-G3)/J3,2)"
        assert sheet["L3"].value.startswith("=MIN(J3,MAX(0,")
        assert sheet["M3"].value == "=IF(L3>=J3,F3,K3*L3)"
        assert sheet["N3"].value == '=IF(L3>=J3,"",K3)'
        assert sheet["Q3"].value == '=IF(L3>=J3,"",F3-M3)'
        assert sheet["C5"].value == "合计"
        assert sheet["F5"].value == "=SUM(F3:F4)"
        assert sheet["K5"].value == "=SUM(K3:K4)"
        assert sheet["M5"].value == "=SUM(M3:M4)"
        assert sheet["N5"].value == "=SUM(N3:N4)"
        assert sheet["O5"].value == "=SUM(O3:O4)"
        assert sheet["Q5"].value == "=SUM(Q3:Q4)"
        assert sheet.freeze_panes == "C3"
    finally:
        workbook.close()


def test_export_adds_single_hierarchical_category_view() -> None:
    workbook = load_workbook(BytesIO(generate_amortization_template()))
    sheet = workbook[AMORTIZATION_SHEET]
    for primary, name, expense, value in (
        ("B类", "资产1", "A费用", 1200),
        ("A类", "资产2", "B费用", 2400),
        ("C类", "资产3", "A费用", 3600),
    ):
        sheet.append(
            [
                primary,
                name,
                expense,
                value,
                0,
                date(2026, 4, 30),
                date(2026, 4, 30),
                12,
            ]
        )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    result = export_amortization_submission(
        "classified.xlsx", stream.getvalue(), NaturalMonth(2026, 6)
    )
    assert result.errors == ()
    assert result.output is not None
    exported = load_workbook(BytesIO(result.output.workbook_bytes), data_only=False)
    try:
        assert exported.sheetnames == ["摊销明细", "分类明细"]
        assert exported.active.title == "摊销明细"
        classified_sheet = exported["分类明细"]
        assert [classified_sheet.cell(row, 1).value for row in range(4, 10)] == [
            "A费用",
            None,
            None,
            "B费用",
            None,
            "总计",
        ]
        assert [classified_sheet.cell(row, 2).value for row in range(4, 10)] == [
            None,
            "B类",
            "C类",
            None,
            "A类",
            None,
        ]
        assert classified_sheet["C4"].value == (
            "=SUMIF('摊销明细'!$E$3:$E$5,A4,'摊销明细'!$N$3:$N$5)"
        )
        assert classified_sheet["C5"].value == (
            "=SUMIFS('摊销明细'!$N$3:$N$5,'摊销明细'!$E$3:$E$5,$A$4,"
            "'摊销明细'!$C$3:$C$5,B5)"
        )
        assert classified_sheet["C9"].value == (
            "=SUM('摊销明细'!$N$3:$N$5)"
        )
    finally:
        exported.close()


def test_negative_correction_is_supported_and_future_start_is_rejected() -> None:
    workbook = load_workbook(BytesIO(generate_amortization_template()))
    sheet = workbook[AMORTIZATION_SHEET]
    sheet.append(
        [
            "软件",
            "发票冲回",
            "管理费用",
            -8962.26,
            0,
            date(2022, 10, 31),
            date(2023, 11, 30),
            60,
        ]
    )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    calculated = calculate_amortization_submission(
        "correction.xlsx", stream.getvalue(), NaturalMonth(2026, 6)
    )
    assert calculated.errors == ()
    assert calculated.calculation is not None
    row = calculated.calculation.rows[0]
    assert row.monthly_amortization == Decimal("-149.37")
    assert row.cumulative_months == 45
    assert row.cumulative_amortization == Decimal("-6721.65")
    assert row.ending_net_value == Decimal("-2240.61")

    future = load_workbook(BytesIO(generate_amortization_template()))
    future_sheet = future[AMORTIZATION_SHEET]
    future_sheet.append(
        [
            "软件",
            "未来开始",
            "管理费用",
            1000,
            0,
            date(2026, 7, 1),
            date(2026, 6, 1),
            12,
        ]
    )
    future_stream = BytesIO()
    future.save(future_stream)
    future.close()
    rejected = calculate_amortization_submission(
        "future.xlsx", future_stream.getvalue(), NaturalMonth(2026, 6)
    )
    assert rejected.calculation is None
    assert {error.error_code for error in rejected.errors} == {
        "START_MONTH_AFTER_CALCULATION"
    }


def test_input_formulas_and_invalid_term_are_rejected() -> None:
    workbook = load_workbook(BytesIO(generate_amortization_template()))
    sheet = workbook[AMORTIZATION_SHEET]
    sheet.append(
        [
            "软件",
            "错误数据",
            "管理费用",
            "=1000+1",
            0,
            date(2026, 1, 1),
            date(2026, 1, 1),
            0,
        ]
    )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    rejected = calculate_amortization_submission(
        "invalid.xlsx", stream.getvalue(), NaturalMonth(2026, 6)
    )
    assert rejected.calculation is None
    assert {error.error_code for error in rejected.errors} == {
        "AMORTIZATION_TERM_INVALID",
        "FORMULA_NOT_ALLOWED",
        "ORIGINAL_VALUE_INVALID",
    }
