from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from loan_interest_accrual.workbook import (
    CURRENCY_NUMBER_FORMAT,
    DATE_NUMBER_FORMAT,
    LOAN_SHEET,
    LOAN_TEMPLATE_HEADERS,
    MAX_LOAN_ROWS,
    MAX_MOVEMENT_ROWS,
    MAX_UPLOAD_BYTES,
    MOVEMENT_SHEET,
    MOVEMENT_TEMPLATE_HEADERS,
    PERCENT_NUMBER_FORMAT,
    generate_standard_template,
)


def test_limits_match_approved_conservative_boundaries() -> None:
    assert MAX_UPLOAD_BYTES == 20 * 1024 * 1024
    assert MAX_LOAN_ROWS == 10_000
    assert MAX_MOVEMENT_ROWS == 100_000


def test_standard_template_has_exact_schema_instructions_and_formats() -> None:
    payload = generate_standard_template()
    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert workbook.sheetnames == [LOAN_SHEET]
    loan_sheet = workbook[LOAN_SHEET]
    assert tuple(cell.value for cell in loan_sheet[1]) == LOAN_TEMPLATE_HEADERS
    assert all(cell.comment and "填写" in cell.comment.text for cell in loan_sheet[1])
    assert loan_sheet.freeze_panes == "A2"
    header_columns = {
        cell.value: cell.column_letter for cell in loan_sheet[1]
    }
    assert loan_sheet[f"{header_columns['期初本金（元）']}2"].number_format == CURRENCY_NUMBER_FORMAT
    assert loan_sheet[f"{header_columns['年利率']}2"].number_format == PERCENT_NUMBER_FORMAT
    assert loan_sheet[f"{header_columns['借款时间']}2"].number_format == DATE_NUMBER_FORMAT

    validations = tuple(loan_sheet.data_validations.dataValidation)
    assert len(validations) == 3
    assert all(not validation.allow_blank for validation in validations)
    assert not any(
        value in {validation.formula1 for validation in validations}
        for value in ('"360,365"', '"是,否"', '"放款,还本"')
    )


def test_standard_template_package_contains_no_active_or_external_content() -> None:
    payload = generate_standard_template()
    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )

    with ZipFile(BytesIO(payload)) as package:
        names = {name.lower() for name in package.namelist()}
    assert not any("vbaproject" in name for name in names)
    assert not any("externallinks" in name for name in names)
    assert "xl/connections.xml" not in names
