from __future__ import annotations

from datetime import date
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from itertools import permutations
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from loan_interest_accrual.domain import (
    DayCountBasis,
    DomainError,
    DomainErrorCode,
    DomainValidationError,
    MovementType,
    NaturalMonth,
    calculate_portfolio,
)
from loan_interest_accrual.workbook import (
    LOAN_SHEET,
    LOAN_TEMPLATE_HEADERS,
    MAX_LOAN_ROWS,
    MAX_MOVEMENT_ROWS,
    MAX_UPLOAD_BYTES,
    MOVEMENT_SHEET,
    MOVEMENT_TEMPLATE_HEADERS,
    WorkbookErrorCode,
    import_workbook,
)
from loan_interest_accrual.workbook import validation as workbook_validation


PERIOD = NaturalMonth(2025, 6)


def loan_row(
    loan_id: object = "L-001",
    company: object = "甲公司",
    contract: object = "HT-001",
    bank: object = "甲银行",
    principal: object = 1000,
    rate: object = 0.025,
    basis: object = 365,
    start: object = date(2025, 1, 1),
    end: object = date(2025, 12, 31),
    capitalized: object = "否",
    note: object = None,
) -> list[object]:
    return [
        company,
        contract,
        bank,
        principal,
        rate,
        start,
        note,
    ]


def movement_row(
    loan_id: object = "L-001",
    event_date: object = date(2025, 6, 15),
    movement_type: object = "放款",
    amount: object = 1,
    note: object = None,
) -> list[object]:
    return [loan_id, event_date, movement_type, amount, note]


def workbook_bytes(
    *,
    loan_headers: tuple[str, ...] = LOAN_TEMPLATE_HEADERS,
    movement_headers: tuple[str, ...] = MOVEMENT_TEMPLATE_HEADERS,
    loans: list[list[object]] | None = None,
    movements: list[list[object]] | None = None,
    include_loan_sheet: bool = True,
    include_movement_sheet: bool = True,
    extra_sheet: bool = False,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    if include_loan_sheet:
        sheet = workbook.create_sheet(LOAN_SHEET)
        sheet.append(list(loan_headers))
        for values in loans if loans is not None else [loan_row()]:
            sheet.append(values)
        if "年利率" in loan_headers:
            rate_column = loan_headers.index("年利率") + 1
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, rate_column).number_format = "0.00%"
        for header in ("借款时间",):
            if header in loan_headers:
                column = loan_headers.index(header) + 1
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "yyyy-mm-dd"
    if include_movement_sheet:
        sheet = workbook.create_sheet(MOVEMENT_SHEET)
        sheet.append(list(movement_headers))
        for values in movements or []:
            sheet.append(values)
        if "变动日期" in movement_headers:
            column = movement_headers.index("变动日期") + 1
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column).number_format = "yyyy-mm-dd"
    if extra_sheet:
        workbook.create_sheet("其他信息")["A1"] = "不得作为输入来源"
    if not workbook.worksheets:
        workbook.create_sheet("其他信息")["A1"] = "非标准工作簿"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def edit_workbook(payload: bytes, edit) -> bytes:
    workbook = load_workbook(BytesIO(payload))
    edit(workbook)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def rewrite_package(
    payload: bytes,
    *,
    replacements: dict[str, bytes] | None = None,
    additions: dict[str, bytes] | None = None,
) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(payload)) as source, ZipFile(
        output, "w", ZIP_DEFLATED
    ) as target:
        for info in source.infolist():
            target.writestr(
                info,
                (replacements or {}).get(info.filename, source.read(info.filename)),
            )
        for name, content in (additions or {}).items():
            target.writestr(name, content)
    return output.getvalue()


def codes(result) -> list[WorkbookErrorCode]:
    return [error.error_code for error in result.errors]


def test_valid_workbook_parses_exact_ids_units_rates_and_domain_types() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row("L-001", company="同公司", contract="同合同"),
            loan_row("L-002", company="同公司", contract="同合同", basis=360),
        ],
        movements=[
            movement_row("L-002", amount=1),
            movement_row("L-002", event_date=date(2025, 6, 15), amount=2),
        ],
        extra_sheet=True,
    )

    result = import_workbook("input.XLSX", payload, PERIOD)

    assert result.errors == ()
    assert result.calculable_input is not None
    assert [loan.loan_id for loan in result.calculable_input.loans] == [
        "ROW-000002",
        "ROW-000003",
    ]
    assert result.calculable_input.loans[0].annual_rate == Decimal("0.025")
    assert result.calculable_input.loans[1].day_count_basis is DayCountBasis.DAYS_360
    assert result.calculable_input.movements == ()


@pytest.mark.parametrize("filename", ["input.xls", "input.xlsm", "input.csv", "input"])
def test_only_xlsx_extension_is_accepted(filename: str) -> None:
    result = import_workbook(filename, workbook_bytes(), PERIOD)
    assert codes(result) == [WorkbookErrorCode.FILE_EXTENSION_INVALID]
    assert result.calculable_input is None


def test_corrupt_xlsx_fails_structurally() -> None:
    result = import_workbook("input.xlsx", b"not-a-zip", PERIOD)
    assert codes(result) == [WorkbookErrorCode.WORKBOOK_OPEN_FAILED]


def test_file_size_limit_is_inclusive_and_exceedance_is_not_opened() -> None:
    at_limit = import_workbook("input.xlsx", b"x" * MAX_UPLOAD_BYTES, PERIOD)
    assert WorkbookErrorCode.FILE_TOO_LARGE not in codes(at_limit)
    assert WorkbookErrorCode.WORKBOOK_OPEN_FAILED in codes(at_limit)

    over_limit = import_workbook("input.xlsx", b"x" * (MAX_UPLOAD_BYTES + 1), PERIOD)
    assert codes(over_limit) == [WorkbookErrorCode.FILE_TOO_LARGE]


@pytest.mark.parametrize(
    ("addition", "expected"),
    [
        ({"xl/vbaProject.bin": b"macro"}, WorkbookErrorCode.MACRO_NOT_ALLOWED),
        (
            {"xl/externalLinks/externalLink1.xml": b"<externalLink/>"},
            WorkbookErrorCode.EXTERNAL_LINK_NOT_ALLOWED,
        ),
        ({"xl/connections.xml": b"<connections/>"}, WorkbookErrorCode.EXTERNAL_LINK_NOT_ALLOWED),
    ],
)
def test_active_and_external_package_content_is_rejected(
    addition: dict[str, bytes],
    expected: WorkbookErrorCode,
) -> None:
    payload = rewrite_package(workbook_bytes(), additions=addition)
    result = import_workbook("input.xlsx", payload, PERIOD)
    assert expected in codes(result)
    assert result.calculable_input is None


def test_duplicate_required_sheet_name_is_rejected() -> None:
    payload = workbook_bytes()
    with ZipFile(BytesIO(payload)) as package:
        root = ElementTree.fromstring(package.read("xl/workbook.xml"))
    namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets = root.find("m:sheets", namespace)
    assert sheets is not None
    sheets[1].set("name", LOAN_SHEET)
    changed = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
    payload = rewrite_package(payload, replacements={"xl/workbook.xml": changed})

    result = import_workbook("input.xlsx", payload, PERIOD)
    assert codes(result) == [WorkbookErrorCode.SHEET_DUPLICATE]


def test_missing_sheets_columns_duplicate_columns_and_aliases_are_not_inferred() -> None:
    missing_sheets = import_workbook(
        "input.xlsx",
        workbook_bytes(include_loan_sheet=False, include_movement_sheet=False),
        PERIOD,
    )
    assert codes(missing_sheets) == [WorkbookErrorCode.SHEET_MISSING]

    alias_headers = tuple(
        "Borrowing Date" if header == "借款时间" else header
        for header in LOAN_TEMPLATE_HEADERS
    )
    missing_column = import_workbook(
        "input.xlsx",
        workbook_bytes(loan_headers=alias_headers),
        PERIOD,
    )
    assert WorkbookErrorCode.COLUMN_MISSING in codes(missing_column)

    duplicate_headers = LOAN_TEMPLATE_HEADERS + ("贷款银行",)
    duplicate_column = import_workbook(
        "input.xlsx",
        workbook_bytes(
            loan_headers=duplicate_headers,
            loans=[loan_row() + ["重复银行"]],
        ),
        PERIOD,
    )
    assert WorkbookErrorCode.COLUMN_DUPLICATE in codes(duplicate_column)


def test_header_errors_do_not_hide_detectable_formula_errors() -> None:
    loan_headers = tuple(
        header for header in LOAN_TEMPLATE_HEADERS if header != "借款时间"
    )
    loan_values = loan_row()
    loan_values.pop(LOAN_TEMPLATE_HEADERS.index("借款时间"))
    payload = edit_workbook(
        workbook_bytes(
            loan_headers=loan_headers,
            loans=[loan_values],
            movements=[movement_row()],
        ),
        lambda workbook: setattr(workbook[MOVEMENT_SHEET]["D2"], "value", "=1+1"),
    )

    result = import_workbook("input.xlsx", payload, PERIOD)

    assert codes(result) == [WorkbookErrorCode.COLUMN_MISSING]
    assert result.errors[0].sheet == LOAN_SHEET
    assert result.errors[0].column_or_field == "借款时间"
    assert result.calculable_input is None


def test_blank_bank_name_is_optional_and_keeps_the_loan_calculable() -> None:
    result = import_workbook(
        "input.xlsx",
        workbook_bytes(loans=[loan_row(bank=None)]),
        PERIOD,
    )

    assert result.errors == ()
    assert result.calculable_input is not None
    assert result.calculable_input.loans[0].bank_name == ""


def test_formula_cells_are_rejected_with_location() -> None:
    payload = edit_workbook(
        workbook_bytes(),
        lambda workbook: setattr(workbook[LOAN_SHEET]["E2"], "value", "=1+1"),
    )
    result = import_workbook("input.xlsx", payload, PERIOD)

    assert codes(result) == [WorkbookErrorCode.FORMULA_NOT_ALLOWED]
    assert result.errors[0].sheet == LOAN_SHEET
    assert result.errors[0].row == 2
    assert result.errors[0].column_or_field == "年利率"


def test_safe_arithmetic_formula_is_allowed_only_for_opening_principal() -> None:
    payload = edit_workbook(
        workbook_bytes(include_movement_sheet=False),
        lambda workbook: setattr(
            workbook[LOAN_SHEET]["D2"],
            "value",
            "=124098000-12409800",
        ),
    )

    result = import_workbook("input.xlsx", payload, PERIOD)

    assert result.errors == ()
    assert result.calculable_input is not None
    assert result.calculable_input.loans[0].opening_principal == Decimal("111688200")

    unsupported = edit_workbook(
        workbook_bytes(include_movement_sheet=False),
        lambda workbook: setattr(workbook[LOAN_SHEET]["D2"], "value", "=A2"),
    )
    unsupported_result = import_workbook("input.xlsx", unsupported, PERIOD)
    assert codes(unsupported_result) == [WorkbookErrorCode.FORMULA_NOT_ALLOWED]


def test_blank_required_start_date_is_distinct_from_invalid_date() -> None:
    blank = import_workbook(
        "input.xlsx",
        workbook_bytes(
            loans=[loan_row(start=None)],
            include_movement_sheet=False,
        ),
        PERIOD,
    )
    invalid = import_workbook(
        "input.xlsx",
        workbook_bytes(
            loans=[loan_row(start="2025-06-01")],
            include_movement_sheet=False,
        ),
        PERIOD,
    )

    assert codes(blank) == [WorkbookErrorCode.REQUIRED_VALUE_MISSING]
    assert blank.errors[0].column_or_field == "借款时间"
    assert codes(invalid) == [WorkbookErrorCode.DATE_INVALID]


def test_loan_ids_and_movement_sheet_are_not_required() -> None:
    payload = workbook_bytes(
        loans=[loan_row("L-001"), loan_row("L-001"), loan_row("")],
        movements=[movement_row("L-001")],
    )
    result = import_workbook("input.xlsx", payload, PERIOD)

    assert result.errors == ()
    assert result.calculable_input is not None
    assert len(result.calculable_input.loans) == 3
    assert result.calculable_input.movements == ()


def test_multiple_errors_are_collected_in_deterministic_sheet_row_field_order() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                "",
                company="",
                principal="1000",
                rate=2.5,
                basis=364,
                start="2025-06-01",
                capitalized="可能",
            )
        ],
        movements=[movement_row("UNKNOWN", date(2025, 7, 1), "错误", 0)],
    )
    payload = edit_workbook(
        payload,
        lambda workbook: setattr(
            workbook[LOAN_SHEET]["F2"], "number_format", "General"
        ),
    )
    result = import_workbook("input.xlsx", payload, PERIOD)

    assert [
        (error.sheet, error.row, error.column_or_field, error.error_code)
        for error in result.errors
    ] == [
        (LOAN_SHEET, 2, "公司名称", WorkbookErrorCode.REQUIRED_VALUE_MISSING),
        (LOAN_SHEET, 2, "期初本金（元）", WorkbookErrorCode.VALUE_TYPE_INVALID),
        (LOAN_SHEET, 2, "年利率", WorkbookErrorCode.INTEREST_RATE_INVALID),
        (LOAN_SHEET, 2, "借款时间", WorkbookErrorCode.DATE_INVALID),
    ]
    assert result.calculable_input is None


def test_date_relationship_rate_basis_enum_and_amount_boundaries() -> None:
    cases = [
        (
            workbook_bytes(
                loans=[
                    loan_row(
                        start=date(2025, 6, 20),
                        end=date(2025, 12, 31),
                        note="15号归还本金",
                    )
                ]
            ),
            WorkbookErrorCode.DATE_RANGE_INVALID,
        ),
        (
            workbook_bytes(loans=[loan_row(rate=0)]),
            WorkbookErrorCode.INTEREST_RATE_INVALID,
        ),
        (
            workbook_bytes(loans=[loan_row(rate=1)]),
            WorkbookErrorCode.INTEREST_RATE_INVALID,
        ),
        (
            workbook_bytes(loans=[loan_row(rate=2.5)]),
            WorkbookErrorCode.INTEREST_RATE_INVALID,
        ),
    ]
    for payload, expected in cases:
        result = import_workbook("input.xlsx", payload, PERIOD)
        assert expected in codes(result)
        assert result.calculable_input is None

    non_percent = edit_workbook(
        workbook_bytes(loans=[loan_row(rate=2.5)]),
        lambda workbook: setattr(
            workbook[LOAN_SHEET]["F2"], "number_format", "General"
        ),
    )
    assert WorkbookErrorCode.INTEREST_RATE_INVALID in codes(
        import_workbook("input.xlsx", non_percent, PERIOD)
    )


def test_rows_outside_selected_month_are_silently_filtered() -> None:
    period = NaturalMonth(2026, 3)
    payload = workbook_bytes(
        loans=[
            loan_row(bank="尚未起息", start=date(2026, 4, 1), rate="无需校验"),
            loan_row(
                bank="此前已还清",
                start=date(2025, 6, 22),
                note="25.9.11还本金",
                principal="无需校验",
            ),
            loan_row(
                bank="仅部分还款",
                start=date(2025, 12, 22),
                note="25.12.19还10%",
            ),
            loan_row(bank="本月有效", start=date(2025, 12, 22), note=None),
        ],
        include_movement_sheet=False,
    )

    imported = import_workbook("input.xlsx", payload, period)

    assert imported.errors == ()
    assert imported.calculable_input is not None
    assert [loan.bank_name for loan in imported.calculable_input.loans] == [
        "仅部分还款",
        "本月有效",
    ]


@pytest.mark.parametrize(
    ("start", "ignored_end", "note", "expected_days"),
    [
        (date(2024, 1, 1), date(2024, 12, 31), None, 30),
        (date(2025, 1, 1), date(2025, 6, 21), None, 30),
        (date(2025, 1, 1), date(2025, 6, 27), None, 30),
        (date(2025, 6, 10), date(2025, 6, 12), None, 20),
        (date(2025, 6, 10), date(2025, 6, 12), "15号归还本金", 5),
        (date(2025, 1, 1), "不是日期", None, 30),
    ],
)
def test_interest_end_date_is_ignored_for_day_count(
    start: date,
    ignored_end: object,
    note: str | None,
    expected_days: int,
) -> None:
    payload = workbook_bytes(
        loans=[loan_row(start=start, end=ignored_end, note=note)],
        include_movement_sheet=False,
    )

    imported = import_workbook("input.xlsx", payload, PERIOD)

    assert imported.errors == ()
    assert imported.calculable_input is not None
    calculation = calculate_portfolio(
        PERIOD,
        imported.calculable_input.loans,
        imported.calculable_input.movements,
    )
    assert calculation.loan_results[0].interest_days == expected_days


def test_negative_principal_uses_domain_calculator_and_locates_movement() -> None:
    payload = workbook_bytes(
        loans=[loan_row(principal=100)],
        movements=[movement_row(movement_type="还本", amount=101)],
    )
    result = import_workbook("input.xlsx", payload, PERIOD)

    assert result.errors == ()
    assert result.calculable_input is not None
    assert result.calculable_input.movements == ()


def test_portfolio_business_validation_errors_are_mapped_to_source_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_portfolio(period, loans, movements):
        assert period is PERIOD
        assert tuple(loan.loan_id for loan in loans) == ("ROW-000002",)
        assert tuple(movements) == ()
        raise DomainValidationError(
            DomainError(
                error_code=DomainErrorCode.NEGATIVE_PRINCIPAL,
                column_or_field="principal",
                message="principal becomes negative at an effective event boundary",
                loan_id="ROW-000002",
            )
        )

    monkeypatch.setattr(workbook_validation, "calculate_portfolio", fail_portfolio)
    payload = workbook_bytes(movements=[movement_row(movement_type="还本")])

    result = import_workbook("input.xlsx", payload, PERIOD)

    assert codes(result) == [WorkbookErrorCode.NEGATIVE_PRINCIPAL]
    assert result.errors[0].sheet == MOVEMENT_SHEET
    assert result.errors[0].row is None
    assert result.calculable_input is None


def test_same_day_movement_order_is_preserved_as_input_but_calculates_by_loan_id() -> None:
    rows = [
        movement_row(amount=10),
        movement_row(movement_type="还本", amount=3),
        movement_row(amount=2),
    ]
    endings = set()
    for ordering in permutations(rows):
        result = import_workbook(
            "input.xlsx",
            workbook_bytes(movements=list(ordering)),
            PERIOD,
        )
        assert result.errors == ()
        assert result.calculable_input is not None
        endings.add(
            tuple(
                (movement.movement_type, movement.amount)
                for movement in result.calculable_input.movements
            )
        )
    assert endings == {()}


def test_row_limits_allow_boundary_and_reject_exceedance_without_truncation() -> None:
    boundary = load_workbook(BytesIO(workbook_bytes()))
    boundary[LOAN_SHEET].cell(MAX_LOAN_ROWS + 1, 1, "L-LAST")
    for column, value in enumerate(loan_row("L-LAST"), start=1):
        boundary[LOAN_SHEET].cell(MAX_LOAN_ROWS + 1, column, value)
    boundary[LOAN_SHEET].cell(MAX_LOAN_ROWS + 1, 6).number_format = "0.00%"
    for column, value in enumerate(movement_row(), start=1):
        boundary[MOVEMENT_SHEET].cell(MAX_MOVEMENT_ROWS + 1, column, value)
    stream = BytesIO()
    boundary.save(stream)
    boundary_result = import_workbook("input.xlsx", stream.getvalue(), PERIOD)
    assert WorkbookErrorCode.LOAN_ROW_LIMIT_EXCEEDED not in codes(boundary_result)
    assert WorkbookErrorCode.MOVEMENT_ROW_LIMIT_EXCEEDED not in codes(boundary_result)

    loan_over = edit_workbook(
        workbook_bytes(),
        lambda workbook: workbook[LOAN_SHEET].cell(
            MAX_LOAN_ROWS + 2, 1, "OVER"
        ),
    )
    assert codes(import_workbook("input.xlsx", loan_over, PERIOD)) == [
        WorkbookErrorCode.LOAN_ROW_LIMIT_EXCEEDED
    ]

    movement_over = edit_workbook(
        workbook_bytes(),
        lambda workbook: workbook[MOVEMENT_SHEET].cell(
            MAX_MOVEMENT_ROWS + 2, 1, "OVER"
        ),
    )
    assert codes(import_workbook("input.xlsx", movement_over, PERIOD)) == []


def test_optional_fields_may_be_blank_and_blank_end_defaults_to_month_end() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                company="甲公司",
                contract=None,
                start=date(2025, 6, 10),
                end=None,
                note=None,
            )
        ],
        include_movement_sheet=False,
    )

    result = import_workbook("input.xlsx", payload, PERIOD)

    assert result.errors == ()
    assert result.calculable_input is not None
    loan = result.calculable_input.loans[0]
    assert loan.company_name == "甲公司"
    assert loan.contract_number == ""
    assert loan.accrual_end == date(2025, 6, 30)


@pytest.mark.parametrize(
    ("start", "note", "expected_days"),
    [
        (date(2025, 1, 1), None, 30),
        (date(2025, 1, 1), "15号归还本金", 15),
        (date(2025, 6, 10), None, 20),
        (date(2025, 6, 10), "15号归还本金", 5),
        (date(2025, 6, 30), None, 0),
        (date(2025, 1, 1), "2025年6月20日归还本金", 20),
        (date(2025, 1, 1), "归还本金日期25号", 25),
        (date(2025, 1, 1), "25.6.27还本金", 27),
        (date(2025, 1, 1), "2025.6.27还本金", 27),
        (date(2025, 1, 1), "25-6-27换本金", 27),
    ],
)
def test_interest_days_follow_start_and_repayment_note_rules(
    start: date,
    note: str | None,
    expected_days: int,
) -> None:
    payload = workbook_bytes(
        loans=[loan_row(start=start, end=None, note=note)],
        include_movement_sheet=False,
    )

    imported = import_workbook("input.xlsx", payload, PERIOD)

    assert imported.errors == ()
    assert imported.calculable_input is not None
    calculation = calculate_portfolio(
        PERIOD,
        imported.calculable_input.loans,
        imported.calculable_input.movements,
    )
    assert calculation.loan_results[0].interest_days == expected_days


def test_partial_repayment_note_reduces_principal_after_repayment_day() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                principal=17_000_000,
                rate=0.025,
                start=date(2025, 1, 1),
                end=None,
                note="25.6.9还700万",
            )
        ],
        include_movement_sheet=False,
    )

    imported = import_workbook("input.xlsx", payload, PERIOD)

    assert imported.errors == ()
    assert imported.calculable_input is not None
    assert len(imported.calculable_input.movements) == 1
    movement = imported.calculable_input.movements[0]
    assert movement.event_date == date(2025, 6, 9)
    assert movement.amount == Decimal("7000000")
    calculation = calculate_portfolio(
        PERIOD,
        imported.calculable_input.loans,
        imported.calculable_input.movements,
    )
    result = calculation.loan_results[0]
    assert [
        (segment.principal, segment.days) for segment in result.segments
    ] == [
        (Decimal("17000000"), 9),
        (Decimal("10000000"), 21),
    ]
    assert result.total_repayments == Decimal("7000000")
    assert result.ending_principal == Decimal("10000000")
    assert result.accrued_interest == Decimal("25208.33")


def test_partial_repayment_before_range_uses_reduced_opening_principal() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                principal=17_000_000,
                rate=0.025,
                start=date(2025, 1, 1),
                end=None,
                note="25.5.9已还700万元",
            )
        ],
        include_movement_sheet=False,
    )

    imported = import_workbook("input.xlsx", payload, PERIOD)

    assert imported.errors == ()
    assert imported.calculable_input is not None
    assert imported.calculable_input.movements == ()
    assert imported.calculable_input.loans[0].opening_principal == Decimal(
        "10000000"
    )


def test_invalid_repayment_day_in_note_is_reported() -> None:
    payload = workbook_bytes(
        loans=[loan_row(end=None, note="31号归还本金")],
        include_movement_sheet=False,
    )

    result = import_workbook("input.xlsx", payload, NaturalMonth(2025, 2))

    assert [(error.column_or_field, error.error_code) for error in result.errors] == [
        ("备注", WorkbookErrorCode.DATE_INVALID)
    ]


def test_source_file_bytes_and_hash_are_never_modified(tmp_path) -> None:
    payload = workbook_bytes(movements=[movement_row()])
    source_path = tmp_path / "source.xlsx"
    source_path.write_bytes(payload)
    before = sha256(source_path.read_bytes()).hexdigest()

    result = import_workbook(source_path.name, source_path.read_bytes(), PERIOD)

    assert result.errors == ()
    assert result.source_bytes == payload
    assert result.source_sha256 == before
    assert source_path.read_bytes() == payload
    assert sha256(source_path.read_bytes()).hexdigest() == before
