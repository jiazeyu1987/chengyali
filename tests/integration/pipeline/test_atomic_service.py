from __future__ import annotations

from datetime import date
from hashlib import sha256
from io import BytesIO

from openpyxl import load_workbook

from loan_interest_accrual.application import (
    calculate_submission,
    export_submission,
)
from loan_interest_accrual.domain import NaturalMonth
from loan_interest_accrual.workbook.export_schema import RESULT_SHEET

from tests.fixtures.export.standard_workbooks import (
    loan_row,
    movement_row,
    workbook_bytes,
)


PERIOD = NaturalMonth(2025, 6)


def _result_principals(output_bytes: bytes) -> list[object]:
    workbook = load_workbook(BytesIO(output_bytes), data_only=True)
    try:
        sheet = workbook[RESULT_SHEET]
        headers = [cell.value for cell in sheet[3]]
        sequence_column = headers.index("序号") + 1
        principal_column = headers.index("期初本金（元）") + 1
        return [
            sheet.cell(row, principal_column).value
            for row in range(4, sheet.max_row + 1)
            if sheet.cell(row, sequence_column).value != "合计"
        ]
    finally:
        workbook.close()


def test_invalid_workbook_is_atomic_and_export_returns_errors_only() -> None:
    payload = workbook_bytes(
        loans=[loan_row("", company="", bank="", rate=2.5)],
        movements=[movement_row("UNKNOWN", event_date=date(2025, 7, 1), amount=0)],
    )
    source_hash = sha256(payload).hexdigest()

    calculated = calculate_submission("invalid.xlsx", payload, PERIOD)
    exported = export_submission("invalid.xlsx", payload, PERIOD)

    assert calculated.source_sha256 == source_hash
    assert calculated.calculation is None
    assert calculated.errors
    assert {error.error_code for error in calculated.errors} == {
        "REQUIRED_VALUE_MISSING",
        "INTEREST_RATE_INVALID",
    }
    assert exported.output is None
    assert exported.errors == calculated.errors


def test_valid_submission_returns_complete_preview_and_preserves_source_hash(
    tmp_path,
) -> None:
    payload = workbook_bytes(
        loans=[
            loan_row("L-CAP", company="甲公司", contract="HT-CAP", capitalized="是"),
            loan_row(
                "L-EXP",
                company="甲公司",
                contract="HT-EXP",
                principal=2000,
                rate=0.036,
                basis=360,
                capitalized="否",
            ),
        ],
        movements=[
            movement_row("L-CAP", date(2025, 6, 15), "放款", 100),
            movement_row("L-CAP", date(2025, 6, 20), "还本", 50),
        ],
    )
    source_path = tmp_path / "source.xlsx"
    source_path.write_bytes(payload)
    before_hash = sha256(source_path.read_bytes()).hexdigest()

    result = calculate_submission(source_path.name, source_path.read_bytes(), PERIOD)
    exported = export_submission(source_path.name, source_path.read_bytes(), PERIOD)

    assert result.errors == ()
    assert result.calculation is not None
    assert result.calculation.source_bytes == payload
    assert result.calculation.source_sha256 == before_hash
    assert source_path.read_bytes() == payload
    assert sha256(source_path.read_bytes()).hexdigest() == before_hash

    rows = result.calculation.loan_rows
    assert [row.loan_id for row in rows] == ["ROW-000002", "ROW-000003"]
    assert rows[0].company_name == "甲公司"
    assert rows[0].contract_number == "HT-CAP"
    assert rows[0].ending_principal == rows[0].opening_principal
    assert rows[0].interest_days == 30
    assert rows[0].capitalized_interest == 0
    assert rows[0].expensed_interest == rows[0].accrued_interest
    assert rows[1].expensed_interest == rows[1].accrued_interest
    assert all(check.status == "通过" for check in result.calculation.checks)

    assert exported.errors == ()
    assert exported.output is not None
    assert exported.output.workbook_bytes != payload
    assert exported.output.calculation.source_sha256 == before_hash


def test_export_recomputes_from_current_bytes_instead_of_prior_state() -> None:
    first_payload = workbook_bytes(loans=[loan_row("L-OLD", company="甲公司")])
    second_payload = workbook_bytes(
        loans=[loan_row("L-NEW", company="乙公司", principal=3000)]
    )

    first_result = calculate_submission("old.xlsx", first_payload, PERIOD)
    exported = export_submission("new.xlsx", second_payload, PERIOD)

    assert first_result.errors == ()
    assert first_result.calculation is not None
    assert [row.loan_id for row in first_result.calculation.loan_rows] == ["ROW-000002"]
    assert exported.errors == ()
    assert exported.output is not None
    assert exported.output.calculation.source_sha256 == sha256(second_payload).hexdigest()
    assert _result_principals(exported.output.workbook_bytes) == [3000]
