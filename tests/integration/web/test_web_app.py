from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from starlette.routing import Mount

from loan_interest_accrual.application.results import ApplicationError
from loan_interest_accrual.domain.errors import DomainErrorCode
from loan_interest_accrual.web.http_models import (
    ERROR_MESSAGE_BY_CODE,
    application_error_to_http,
)
from loan_interest_accrual.workbook.schema import WorkbookErrorCode
from loan_interest_accrual.web import create_app
from tests.fixtures.web.workbooks import loan_row, movement_row, workbook_bytes


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
EXPECTED_EXPORT_SHEETS = ["计提结果"]
WEB_ROOT = (
    Path(__file__).resolve().parents[3]
    / "src"
    / "loan_interest_accrual"
    / "web"
)
USER_VISIBLE_ENGLISH_DIAGNOSTICS = (
    "must be",
    "not found",
    "greater than",
    "inside the selected month",
)
HTTP_ONLY_ERROR_CODES = {
    "PERIOD_REQUIRED",
    "FILE_REQUIRED",
    "CALCULABLE_INPUT_MISSING",
    "EXPORT_CHECK_FAILED",
}


def _client() -> TestClient:
    return TestClient(create_app())


def _xlsx_upload(filename: str, payload: bytes) -> dict[str, tuple[str, bytes, str]]:
    return {"file": (filename, payload, XLSX_MEDIA_TYPE)}


def _attachment_filename(content_disposition: str) -> str:
    marker = "filename*=UTF-8''"
    assert marker in content_disposition
    return unquote(content_disposition.split(marker, 1)[1])


def _historical_workbook_upload() -> tuple[str, bytes]:
    matches = sorted(Path("doc").glob("*.xlsx"))
    assert len(matches) == 1
    path = matches[0]
    return path.name, path.read_bytes()


def _assert_simplified_chinese_error_messages(
    errors: list[dict[str, object]],
) -> None:
    assert errors
    for error in errors:
        message = str(error["message"])
        assert re.search(r"[\u4e00-\u9fff]", message), error
        normalized = message.lower()
        assert all(
            phrase not in normalized
            for phrase in USER_VISIBLE_ENGLISH_DIAGNOSTICS
        ), error


def test_http_error_catalog_covers_all_production_codes_and_rejects_unknown() -> None:
    expected_codes = (
        {code.value for code in DomainErrorCode}
        | {code.value for code in WorkbookErrorCode}
        | HTTP_ONLY_ERROR_CODES
    )

    assert set(ERROR_MESSAGE_BY_CODE) == expected_codes
    _assert_simplified_chinese_error_messages(
        [
            {"error_code": code, "message": message}
            for code, message in ERROR_MESSAGE_BY_CODE.items()
        ]
    )

    with pytest.raises(ValueError, match="UNKNOWN_ERROR"):
        application_error_to_http(
            ApplicationError(
                error_code="UNKNOWN_ERROR",
                sheet=None,
                row=None,
                column_or_field="unknown",
                message="must not be exposed",
            )
        )


def test_homepage_health_and_local_assets_expose_the_browser_workflow() -> None:
    with _client() as client:
        homepage = client.get("/")
        health = client.get("/health")
        stylesheet = client.get("/static/styles.css")
        script = client.get("/static/app.js")
        favicon = client.get("/static/favicon.svg")

    assert homepage.status_code == 200
    assert homepage.headers["content-type"].startswith("text/html")
    assert "贷款利息自动计提" in homepage.text
    assert homepage.text.count('type="date"') == 2
    assert 'accept=".xlsx"' in homepage.text
    assert 'href="/template"' in homepage.text
    assert "计算并预览" in homepage.text
    assert "导出结果" in homepage.text
    assert 'src="/static/app.js"' in homepage.text
    assert 'href="/static/styles.css"' in homepage.text
    assert 'rel="icon"' in homepage.text
    assert 'href="/static/favicon.svg"' in homepage.text
    assert "http://" not in homepage.text
    assert "https://" not in homepage.text
    assert health.status_code == 200
    assert health.json() == {"status": "ok"}
    assert stylesheet.status_code == 200
    assert script.status_code == 200
    assert favicon.status_code == 200
    assert favicon.headers["content-type"].startswith("image/svg+xml")

    for path in WEB_ROOT.rglob("*"):
        if path.suffix in {".html", ".css", ".js"}:
            text = path.read_text(encoding="utf-8")
            assert "http://" not in text
            assert "https://" not in text
            assert "localStorage" not in text
            assert "sessionStorage" not in text


def test_homepage_uses_the_jinja_template_and_mounted_static_files() -> None:
    application = create_app()

    with TestClient(application) as client:
        homepage = client.get("/")
        stylesheet = client.get("/static/styles.css")
        script = client.get("/static/app.js")

    assert "按所选日期区间校验贷款数据" in homepage.text
    assert any(
        isinstance(route, Mount)
        and route.path == "/static"
        and route.name == "static"
        for route in application.routes
    )
    assert stylesheet.content == (WEB_ROOT / "static" / "styles.css").read_bytes()
    assert script.content == (WEB_ROOT / "static" / "app.js").read_bytes()


def test_template_download_has_exact_filename_media_type_and_sheets() -> None:
    with _client() as client:
        response = client.get("/template")

    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert (
        _attachment_filename(response.headers["content-disposition"])
        == "贷款利息计提输入模板.xlsx"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    try:
        assert workbook.sheetnames == ["贷款主表"]
    finally:
        workbook.close()


def test_calculate_returns_complete_success_preview_and_checks() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                "L-CAP",
                company="甲公司",
                contract="HT-CAP",
                capitalized="是",
            ),
            loan_row(
                "L-EXP",
                company="乙公司",
                contract="HT-EXP",
                principal=2000,
                rate=0.036,
                basis=360,
                capitalized="否",
            ),
        ],
        movements=[
            movement_row("L-CAP", date(2025, 6, 15), "放款", 100),
            movement_row("L-CAP", date(2025, 6, 20), "还本", 50),
        ],
    )

    with _client() as client:
        response = client.post(
            "/calculate",
            data={"calculation_month": "2025-06"},
            files=_xlsx_upload("valid.xlsx", payload),
        )

    assert response.status_code == 200
    body = response.json()
    assert body["success"] is True
    assert body["calculation_month"] == "2025-06"
    assert body["validation_status"] == "通过"
    assert body["source_filename"] == "valid.xlsx"
    assert body["source_sha256"]
    assert body["errors"] == []
    assert len(body["preview"]) == 2
    assert body["preview"] == [
      {
        "sequence": 1,
        "company_name": "甲公司",
        "bank_name": "甲银行",
        "opening_principal": "1000.00",
        "annual_rate": "2.5000%",
        "borrowing_time": "2025-01-01",
        "accrued_interest": "2.08",
        "interest_days": 30,
      },
      {
        "sequence": 2,
        "company_name": "乙公司",
        "bank_name": "甲银行",
        "opening_principal": "2000.00",
        "annual_rate": "3.6000%",
        "borrowing_time": "2025-01-01",
        "accrued_interest": "6.00",
        "interest_days": 30,
      },
    ]
    assert body["company_summaries"] == [
        {
            "company_name": "甲公司",
            "opening_principal": "1000.00",
            "accrued_interest": "2.08",
        },
        {
            "company_name": "乙公司",
            "opening_principal": "2000.00",
            "accrued_interest": "6.00",
        },
    ]
    assert body["checks"]
    assert all(check["status"] == "通过" for check in body["checks"])


def test_calculate_shows_filled_bank_and_leaves_optional_bank_blank() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(company="甲公司", bank="甲银行", principal=1000),
            loan_row(company="乙公司", bank=None, principal=2000),
        ]
    )

    with _client() as client:
        response = client.post(
            "/calculate",
            data={
                "calculation_start_date": "2025-06-01",
                "calculation_end_date": "2025-06-30",
            },
            files=_xlsx_upload("optional-bank.xlsx", payload),
        )

    assert response.status_code == 200
    body = response.json()
    assert [row["bank_name"] for row in body["preview"]] == ["甲银行", ""]
    assert [row["opening_principal"] for row in body["preview"]] == [
        "1000.00",
        "2000.00",
    ]
    assert body["summary"]["opening_principal"] == "3000.00"


def test_calculate_date_range_intersects_borrowing_and_repayment_dates() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                bank="整段贷款",
                principal=1_000_000,
                rate=0.036,
                start=date(2025, 5, 1),
                note=None,
            ),
            loan_row(
                bank="区间内借还",
                principal=1_000_000,
                rate=0.036,
                start=date(2025, 7, 10),
                note="25.8.15还本金",
            ),
            loan_row(
                bank="区间前已还清",
                principal=1_000_000,
                rate=0.036,
                start=date(2025, 1, 1),
                note="25.5.31还本金",
            ),
        ]
    )

    with _client() as client:
        response = client.post(
            "/calculate",
            data={
                "calculation_start_date": "2025-06-10",
                "calculation_end_date": "2025-08-20",
            },
            files=_xlsx_upload("range.xlsx", payload),
        )

    assert response.status_code == 200
    body = response.json()
    assert body["calculation_month"] == "2025-06-10至2025-08-20"
    assert [
        (row["bank_name"], row["interest_days"], row["accrued_interest"])
        for row in body["preview"]
    ] == [
        ("整段贷款", 72, "7200.00"),
        ("区间内借还", 36, "3600.00"),
    ]


def test_partial_repayment_note_is_split_before_and_after_repayment_day() -> None:
    payload = workbook_bytes(
        loans=[
            loan_row(
                company="杭州唯强医疗科技有限公司",
                bank="测试银行",
                principal=17_000_000,
                rate=0.025,
                start=date(2025, 9, 18),
                note="26.6.9还700万",
            )
        ]
    )

    with _client() as client:
        response = client.post(
            "/calculate",
            data={
                "calculation_start_date": "2026-06-01",
                "calculation_end_date": "2026-06-30",
            },
            files=_xlsx_upload("partial-repayment.xlsx", payload),
        )

    assert response.status_code == 200
    body = response.json()
    assert [
        (
            row["opening_principal"],
            row["interest_days"],
            row["accrued_interest"],
        )
        for row in body["preview"]
    ] == [
        ("17000000.00", 9, "10625.00"),
        ("10000000.00", 21, "14583.33"),
    ]
    assert body["summary"]["loan_count"] == 1
    assert body["summary"]["opening_principal"] == "17000000.00"
    assert body["summary"]["total_repayments"] == "7000000.00"
    assert body["summary"]["ending_principal"] == "10000000.00"
    assert body["summary"]["accrued_interest"] == "25208.33"
    assert body["company_summaries"] == [
        {
            "company_name": "杭州唯强医疗科技有限公司",
            "opening_principal": "17000000.00",
            "accrued_interest": "25208.33",
        }
    ]


def test_calculate_accepts_known_historical_workbook_month() -> None:
    filename, payload = _historical_workbook_upload()

    with _client() as client:
        response = client.post(
            "/calculate",
            data={"calculation_month": "2024-01"},
            files=_xlsx_upload(filename, payload),
        )

    assert response.status_code == 200
    body = response.json()
    assert body["success"] is True
    assert any(
        row["accrued_interest"] == "10763.89" for row in body["preview"]
    )


def test_historical_workbook_missing_selected_month_has_actionable_error() -> None:
    filename, payload = _historical_workbook_upload()

    with _client() as client:
        response = client.post(
            "/calculate",
            data={"calculation_month": "2026-07"},
            files=_xlsx_upload(filename, payload),
        )

    assert response.status_code == 422
    body = response.json()
    assert body["success"] is False
    assert [error["error_code"] for error in body["errors"]] == [
        "HISTORICAL_PERIOD_NOT_FOUND"
    ]
    assert body["errors"][0]["sheet"] == "历史工作簿"


def test_calculate_returns_atomic_structured_errors_without_preview() -> None:
    payload = workbook_bytes(
        loans=[loan_row("", company="", bank="", rate=2.5)],
        movements=[
            movement_row(
                "UNKNOWN",
                event_date=date(2025, 7, 1),
                movement_type="错误",
                amount=0,
            )
        ],
    )

    with _client() as client:
        response = client.post(
            "/calculate",
            data={"calculation_month": "2025-06"},
            files=_xlsx_upload("invalid.xlsx", payload),
        )

    assert response.status_code == 422
    assert "content-disposition" not in response.headers
    body = response.json()
    assert body["success"] is False
    assert body["validation_status"] == "失败"
    assert body["preview"] == []
    assert body["checks"] == []
    assert {error["error_code"] for error in body["errors"]} == {
        "REQUIRED_VALUE_MISSING",
        "INTEREST_RATE_INVALID",
    }
    assert all(
        set(error)
        == {"error_code", "sheet", "row", "column_or_field", "message"}
        for error in body["errors"]
    )
    _assert_simplified_chinese_error_messages(body["errors"])


def test_request_validation_errors_use_the_same_structured_contract() -> None:
    payload = workbook_bytes()
    with _client() as client:
        invalid_month = client.post(
            "/calculate",
            data={"calculation_month": "2025/06"},
            files=_xlsx_upload("valid.xlsx", payload),
        )
        missing_file = client.post(
            "/calculate",
            data={"calculation_month": "2025-06"},
        )

    assert invalid_month.status_code == 422
    assert invalid_month.json()["errors"] == [
        {
            "error_code": "PERIOD_INVALID",
            "sheet": None,
            "row": None,
            "column_or_field": "calculation_month",
                "message": "日期区间必须使用有效的 YYYY-MM-DD 格式，且结束日期不得早于开始日期",
        }
    ]
    assert missing_file.status_code == 422
    assert missing_file.json()["errors"] == [
        {
            "error_code": "FILE_REQUIRED",
            "sheet": None,
            "row": None,
            "column_or_field": "file",
            "message": "请选择一个 .xlsx 文件",
        }
    ]


def test_export_recomputes_current_submission_and_streams_exact_workbook() -> None:
    old_payload = workbook_bytes(loans=[loan_row("L-OLD", company="甲公司")])
    new_payload = workbook_bytes(
        loans=[loan_row("L-NEW", company="乙公司", principal=3000)]
    )

    with _client() as client:
        calculated = client.post(
            "/calculate",
            data={"calculation_month": "2025-06"},
            files=_xlsx_upload("old.xlsx", old_payload),
        )
        exported = client.post(
            "/export",
            data={"calculation_month": "2025-06"},
            files=_xlsx_upload("new.xlsx", new_payload),
        )

    assert calculated.status_code == 200
    assert calculated.json()["preview"][0]["opening_principal"] == "1000.00"
    assert exported.status_code == 200
    assert exported.headers["content-type"] == XLSX_MEDIA_TYPE
    assert (
        _attachment_filename(exported.headers["content-disposition"])
        == "计提结果_2025-06.xlsx"
    )
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    try:
        assert workbook.sheetnames == EXPECTED_EXPORT_SHEETS
        result_sheet = workbook["计提结果"]
        assert result_sheet["A1"].value == "计提区间  2025-06-01 至 2025-06-30"
        headers = [result_sheet.cell(3, column).value for column in range(1, 9)]
        assert headers == [
            "序号",
            "公司名称",
            "贷款银行",
            "期初本金（元）",
            "年利率",
            "借款时间",
            "区间应提利息（元）",
            "区间应提利息天数",
        ]
        principal_column = headers.index("期初本金（元）") + 1
        assert result_sheet.cell(4, principal_column).value == 3000
        assert [result_sheet.cell(3, column).value for column in range(10, 13)] == [
            "公司名称",
            "本金合计（元）",
            "利息合计（元）",
        ]
        assert result_sheet["J4"].value == "乙公司"
        assert result_sheet["K4"].value == 3000
        assert result_sheet["L4"].value == 6.25
    finally:
        workbook.close()


def test_export_failure_returns_json_errors_and_no_download() -> None:
    with _client() as client:
        response = client.post(
            "/export",
            data={"calculation_month": "2025-06"},
            files=_xlsx_upload("invalid.xlsx", b"not-an-xlsx-package"),
        )

    assert response.status_code == 422
    assert response.headers["content-type"].startswith("application/json")
    assert "content-disposition" not in response.headers
    body = response.json()
    assert body["success"] is False
    assert body["preview"] == []
    assert body["errors"][0]["error_code"] == "WORKBOOK_OPEN_FAILED"
    _assert_simplified_chinese_error_messages(body["errors"])
