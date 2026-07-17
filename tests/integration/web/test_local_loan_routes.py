from __future__ import annotations

from io import BytesIO
from urllib.parse import unquote

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from loan_interest_accrual.web import create_app
from tests.fixtures.web.workbooks import workbook_bytes


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _upload(payload: bytes) -> dict[str, tuple[str, bytes, str]]:
    return {"file": ("loan.xlsx", payload, XLSX_MEDIA_TYPE)}


def test_embedded_loan_final_version_calculates_and_exports_without_remote_service() -> None:
    payload = workbook_bytes()
    period = {
        "calculation_start_date": "2025-06-01",
        "calculation_end_date": "2025-06-30",
    }

    with TestClient(create_app()) as client:
        health = client.get("/loan-interest/health")
        template = client.get("/loan-interest/template")
        calculated = client.post(
            "/loan-interest/calculate",
            data=period,
            files=_upload(payload),
        )
        exported = client.post(
            "/loan-interest/export",
            data=period,
            files=_upload(payload),
        )

    assert health.status_code == 200
    assert health.json() == {"status": "ok"}
    assert template.status_code == 200
    assert "贷款利息计提输入模板" in unquote(
        template.headers["content-disposition"]
    )

    assert calculated.status_code == 200
    result = calculated.json()
    assert result["success"] is True
    assert result["calculation_month"] == "2025-06-01至2025-06-30"
    assert result["preview"][0]["company_name"] == "甲公司"
    assert result["validation_status"] == "通过"

    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    try:
        assert workbook.sheetnames == ["计提结果"]
    finally:
        workbook.close()
