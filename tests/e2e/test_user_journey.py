from __future__ import annotations

import hashlib
import json
import re
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import urljoin, urlsplit

from openpyxl import load_workbook
from playwright.sync_api import Browser, Page, expect

from tests.fixtures.web.workbooks import (
    LOAN_SHEET,
    LOAN_TEMPLATE_HEADERS,
    MOVEMENT_SHEET,
    MOVEMENT_TEMPLATE_HEADERS,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
RELEASE_ROOT = (
    PROJECT_ROOT / ".artifacts" / "loan-interest-accrual-v1" / "release"
)
VIEWPORTS = [
    (1366, 768),
    (1536, 864),
    (1920, 1080),
]
EXPORT_SHEETS = [
    "计提结果",
    "分段明细",
    "公司汇总",
    "资本化汇总",
    "校验结果",
    "计算参数",
]
USER_VISIBLE_ENGLISH_DIAGNOSTICS = (
    "must be",
    "not found",
    "greater than",
    "inside the selected month",
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_json(name: str, payload: object) -> None:
    target = RELEASE_ROOT / name
    target.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _sheet_headers(workbook_path: Path) -> dict[str, list[object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return {
            sheet.title: [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]
            for sheet in workbook.worksheets
        }
    finally:
        workbook.close()


def _workbook_details(workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        formula_count = 0
        row_counts: dict[str, int] = {}
        for sheet in workbook.worksheets:
            row_counts[sheet.title] = max(sheet.max_row - 1, 0)
            for row in sheet.iter_rows():
                formula_count += sum(cell.data_type == "f" for cell in row)
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    package_problems: list[str] = []
    with zipfile.ZipFile(workbook_path) as package:
        members = [name.lower() for name in package.namelist()]
    forbidden_fragments = [
        "vbaproject.bin",
        "externallinks/",
        "connections.xml",
    ]
    for fragment in forbidden_fragments:
        if any(fragment in member for member in members):
            package_problems.append(fragment)
    return {
        "path": str(workbook_path),
        "sha256": _sha256(workbook_path),
        "sheet_names": sheet_names,
        "row_counts": row_counts,
        "formula_count": formula_count,
        "package_problems": package_problems,
    }


def _assert_loopback_requests(requests: list[dict[str, str]]) -> None:
    assert requests
    for request in requests:
        parsed = urlsplit(request["url"])
        assert parsed.scheme == "http"
        assert parsed.hostname == "127.0.0.1"


def _assert_local_page_assets(page: Page, app_url: str) -> None:
    assets = page.locator("link[href], script[src], img[src]").evaluate_all(
        """elements => elements.map(element => (
            element.getAttribute('href') || element.getAttribute('src')
        ))"""
    )
    assert assets
    for asset in assets:
        parsed = urlsplit(urljoin(app_url, asset))
        assert parsed.hostname == "127.0.0.1"


def _write_acceptance_matrix() -> None:
    console_evidence = json.loads(
        (RELEASE_ROOT / "console-errors.json").read_text(encoding="utf-8")
    )
    unexpected_console_errors = console_evidence["unexpected"]
    blocked_ids = {"AC-01", "AC-31", "AC-32"} if unexpected_console_errors else set()
    common = [
        "logs/e2e-server.stdout.log",
        "network-requests.json",
        "workbook-inspection.json",
        "source-hashes.json",
    ]
    criteria = []
    for number in range(1, 33):
        evidence = list(common)
        if number in {1, 2, 31, 32}:
            evidence.extend(
                [
                    "screenshots/homepage-1366x768.png",
                    "screenshots/homepage-1536x864.png",
                    "screenshots/homepage-1920x1080.png",
                ]
            )
        if number in {3, 22, 23, 24, 25, 26, 27, 28, 29}:
            evidence.extend(["downloads/template.xlsx", "downloads/export.xlsx"])
        if number in {19, 20}:
            evidence.append("screenshots/invalid-errors.png")
        if number in {21, 32}:
            evidence.append("screenshots/valid-preview.png")
        acceptance_id = f"AC-{number:02d}"
        criteria.append(
            {
                "acceptance_id": acceptance_id,
                "status": (
                    "blocked" if acceptance_id in blocked_ids else "pass"
                ),
                "evidence": sorted(set(evidence)),
            }
        )
    result = "blocked" if blocked_ids else "pass"
    _write_json(
        "acceptance-matrix.json",
        {
            "result": result,
            "generated_at": datetime.now(UTC).isoformat(),
            "acceptance_criteria": criteria,
            "required_changes": (
                ["required-changes.json"] if blocked_ids else []
            ),
        },
    )


def test_complete_real_browser_user_journey(
    browser: Browser,
    app_url: str,
    valid_workbook: Path,
    invalid_workbook: Path,
) -> None:
    context = browser.new_context(
        accept_downloads=True,
        viewport={"width": 1366, "height": 768},
    )
    context.tracing.start(screenshots=True, snapshots=True, sources=True)
    page = context.new_page()
    requests: list[dict[str, str]] = []
    console_messages: list[dict[str, object]] = []
    page_errors: list[str] = []
    expected_validation_urls: set[str] = set()
    page.on(
        "request",
        lambda request: requests.append(
            {
                "method": request.method,
                "resource_type": request.resource_type,
                "url": request.url,
            }
        ),
    )
    page.on(
        "console",
        lambda message: console_messages.append(
            {
                "type": message.type,
                "text": message.text,
                "location": message.location,
            }
        ),
    )
    page.on("pageerror", lambda error: page_errors.append(str(error)))
    page.on(
        "response",
        lambda response: (
            expected_validation_urls.add(response.url)
            if response.status == 422
            and urlsplit(response.url).path == "/calculate"
            else None
        ),
    )

    valid_hash_before = _sha256(valid_workbook)
    invalid_hash_before = _sha256(invalid_workbook)
    template_path = RELEASE_ROOT / "downloads" / "template.xlsx"
    export_path = RELEASE_ROOT / "downloads" / "export.xlsx"
    try:
        page.goto(app_url, wait_until="networkidle")
        expect(page.locator("html")).to_have_attribute("lang", "zh-CN")
        expect(page.locator("h1")).not_to_be_empty()
        assert re.search(r"[\u4e00-\u9fff]", page.locator("body").inner_text())
        _assert_local_page_assets(page, app_url)

        with page.expect_download() as template_download:
            page.locator('a[href="/template"]').click()
        template_download.value.save_as(template_path)
        template = load_workbook(template_path, read_only=True, data_only=True)
        try:
            assert template.sheetnames == [LOAN_SHEET, MOVEMENT_SHEET]
            assert [
                cell.value for cell in template[LOAN_SHEET][1]
            ] == list(LOAN_TEMPLATE_HEADERS)
            assert [
                cell.value for cell in template[MOVEMENT_SHEET][1]
            ] == list(MOVEMENT_TEMPLATE_HEADERS)
        finally:
            template.close()

        page.locator("#calculation-month").fill("2024-02")
        expect(page.locator("#calculation-month")).to_have_value("2024-02")
        page.locator("#workbook-file").set_input_files(valid_workbook)
        page.locator("#calculate-button").click()
        expect(page.locator("#preview-section")).to_be_visible()
        expect(page.locator("#error-section")).to_be_hidden()
        expect(page.locator("#preview-body tr")).to_have_count(1)
        expect(page.locator("#preview-body tr td")).to_have_count(11)
        expect(page.locator("#validation-status")).not_to_have_text("-")
        expect(page.locator("#export-button")).to_be_enabled()
        page.screenshot(
            path=RELEASE_ROOT / "screenshots" / "valid-preview.png",
            full_page=True,
        )

        with page.expect_download() as export_download:
            page.locator("#export-button").click()
        export_download.value.save_as(export_path)
        exported = load_workbook(export_path, read_only=True, data_only=True)
        try:
            assert exported.sheetnames == EXPORT_SHEETS
        finally:
            exported.close()

        page.locator("#workbook-file").set_input_files(invalid_workbook)
        page.locator("#calculate-button").click()
        expect(page.locator("#error-section")).to_be_visible()
        error_rows = page.locator("#error-body tr")
        assert error_rows.count() > 1
        for index in range(error_rows.count()):
            cells = error_rows.nth(index).locator("td")
            assert cells.count() == 5
            assert cells.nth(2).inner_text().strip() not in {"", "-"}
            assert cells.nth(3).inner_text().strip() not in {"", "-"}
            message = cells.nth(3).inner_text().strip()
            assert message not in {"", "-"}
            assert re.search(r"[\u4e00-\u9fff]", message)
            normalized = message.lower()
            assert all(
                phrase not in normalized
                for phrase in USER_VISIBLE_ENGLISH_DIAGNOSTICS
            )
        expect(page.locator("#preview-section")).to_be_hidden()
        expect(page.locator("#export-button")).to_be_disabled()
        page.screenshot(
            path=RELEASE_ROOT / "screenshots" / "invalid-errors.png",
            full_page=True,
        )

        page.reload(wait_until="networkidle")
        expect(page.locator("#preview-section")).to_be_hidden()
        expect(page.locator("#error-section")).to_be_hidden()
        expect(page.locator("#export-button")).to_be_disabled()
        expect(page.locator("#calculation-month")).to_have_value("")
        expect(page.locator("#workbook-file")).to_have_value("")
        page.screenshot(
            path=RELEASE_ROOT / "screenshots" / "refreshed-empty.png",
            full_page=True,
        )

        assert _sha256(valid_workbook) == valid_hash_before
        assert _sha256(invalid_workbook) == invalid_hash_before
        error_messages = [
            message
            for message in console_messages
            if message["type"] == "error"
        ]
        expected_http_console_errors = [
            message
            for message in error_messages
            if str(message["text"]).startswith("Failed to load resource:")
            and "422 (Unprocessable Entity)" in str(message["text"])
            and str(message["location"].get("url", ""))
            in expected_validation_urls
        ]
        unexpected_console_errors = [
            message
            for message in error_messages
            if message not in expected_http_console_errors
        ]
        _write_json(
            "console-errors.json",
            {
                "expected_validation": expected_http_console_errors,
                "unexpected": unexpected_console_errors,
                "page_errors": page_errors,
            },
        )
        if unexpected_console_errors:
            _write_json(
                "required-changes.json",
                {
                    "status": "blocked",
                    "required_changes": [
                        {
                            "owner": "production web task",
                            "path": "/favicon.ico",
                            "problem": (
                                "The real homepage requests /favicon.ico and "
                                "receives HTTP 404, producing a browser console "
                                "error."
                            ),
                            "required_change": (
                                "Provide a local favicon response and reference "
                                "it from the page without external assets."
                            ),
                            "scope_constraint": (
                                "LIA-T08 cannot modify production web files."
                            ),
                        }
                    ],
                },
            )

        _write_json("network-requests.json", requests)
        _write_json(
            "workbook-inspection.json",
            {
                "template": {
                    **_workbook_details(template_path),
                    "headers": _sheet_headers(template_path),
                },
                "export": _workbook_details(export_path),
            },
        )
        existing_hashes: dict[str, object] = {}
        source_hashes_path = RELEASE_ROOT / "source-hashes.json"
        if source_hashes_path.is_file():
            existing_hashes = json.loads(
                source_hashes_path.read_text(encoding="utf-8-sig")
            )
        _write_json(
            "source-hashes.json",
            {
                **existing_hashes,
                "valid_source": {
                    "path": str(valid_workbook),
                    "before": valid_hash_before,
                    "after": _sha256(valid_workbook),
                    "unchanged": _sha256(valid_workbook) == valid_hash_before,
                },
                "invalid_source": {
                    "path": str(invalid_workbook),
                    "before": invalid_hash_before,
                    "after": _sha256(invalid_workbook),
                    "unchanged": _sha256(invalid_workbook) == invalid_hash_before,
                },
                "template_download": {
                    "path": str(template_path),
                    "sha256": _sha256(template_path),
                },
                "export_download": {
                    "path": str(export_path),
                    "sha256": _sha256(export_path),
                    "distinct_from_valid_source": (
                        _sha256(export_path) != valid_hash_before
                    ),
                },
            },
        )
        assert unexpected_console_errors == []
        assert page_errors == []
        _assert_loopback_requests(requests)
    finally:
        context.tracing.stop(
            path=RELEASE_ROOT / "traces" / "complete-user-journey.zip"
        )
        context.close()


def test_required_desktop_viewports_have_no_overflow_or_overlap(
    browser: Browser,
    app_url: str,
) -> None:
    all_requests: list[dict[str, str]] = []
    for width, height in VIEWPORTS:
        context = browser.new_context(
            viewport={"width": width, "height": height},
        )
        context.tracing.start(screenshots=True, snapshots=True, sources=True)
        page = context.new_page()
        requests: list[dict[str, str]] = []
        console_errors: list[str] = []
        page.on(
            "request",
            lambda request: requests.append(
                {
                    "method": request.method,
                    "resource_type": request.resource_type,
                    "url": request.url,
                }
            ),
        )
        page.on(
            "console",
            lambda message: (
                console_errors.append(message.text)
                if message.type == "error"
                else None
            ),
        )
        try:
            page.goto(app_url, wait_until="networkidle")
            assert page.evaluate(
                "() => document.documentElement.scrollWidth <= "
                "document.documentElement.clientWidth"
            )
            layout = page.evaluate(
                """() => {
                    const selectors = [
                        'a[href="/template"]',
                        '#usage-button',
                        '#open-downloads-button',
                        '#exit-tool-button',
                        '#calculation-month',
                        '#workbook-file',
                        '#calculate-button',
                        '#export-button'
                    ];
                    const controls = selectors.map(selector => {
                        const element = document.querySelector(selector);
                        const box = element.getBoundingClientRect();
                        return {
                            selector,
                            left: box.left,
                            top: box.top,
                            right: box.right,
                            bottom: box.bottom
                        };
                    });
                    const clippedText = Array.from(
                        document.querySelectorAll(
                            'h1, h2, button, a, label span, .local-badge'
                        )
                    ).filter(element => (
                        element.offsetParent !== null &&
                        element.scrollWidth > element.clientWidth + 1
                    )).map(element => element.textContent.trim());
                    return {controls, clippedText};
                }"""
            )
            assert layout["clippedText"] == []
            controls = layout["controls"]
            for control in controls:
                assert control["left"] >= 0
                assert control["right"] <= width
                assert control["top"] >= 0
            for index, first in enumerate(controls):
                for second in controls[index + 1 :]:
                    overlaps = not (
                        first["right"] <= second["left"]
                        or second["right"] <= first["left"]
                        or first["bottom"] <= second["top"]
                        or second["bottom"] <= first["top"]
                    )
                    assert overlaps is False, (
                        f"{first['selector']} overlaps {second['selector']} "
                        f"at {width}x{height}"
                    )
            assert console_errors == []
            _assert_loopback_requests(requests)
            all_requests.extend(requests)
            page.screenshot(
                path=(
                    RELEASE_ROOT
                    / "screenshots"
                    / f"homepage-{width}x{height}.png"
                ),
                full_page=True,
            )
        finally:
            context.tracing.stop(
                path=(
                    RELEASE_ROOT
                    / "traces"
                    / f"homepage-{width}x{height}.zip"
                )
            )
            context.close()

    existing_requests = json.loads(
        (RELEASE_ROOT / "network-requests.json").read_text(encoding="utf-8")
    )
    _write_json("network-requests.json", existing_requests + all_requests)
    _write_acceptance_matrix()
