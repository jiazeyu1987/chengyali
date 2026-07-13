import json
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from loan_interest_accrual.domain import NaturalMonth
from loan_interest_accrual.application import calculate_submission, export_submission
from loan_interest_accrual.workbook.export_schema import RESULT_SHEET
from loan_interest_accrual.workbook.importer import import_workbook
from loan_interest_accrual.workbook.schema import WorkbookErrorCode
from tools.historical_validation.validator import (
    ALLOWED_REASON_CODES,
    MANIFEST_PATH,
    run_validation,
)


ARTIFACT_DIR = Path(".artifacts/loan-interest-accrual-v1/historical")


def _historical_workbook_path() -> Path:
    matches = sorted(Path("doc").glob("*.xlsx"))
    assert len(matches) == 1
    return matches[0]


def _sha256_file(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def test_historical_validator_writes_deterministic_read_only_report() -> None:
    source_path = _historical_workbook_path()
    before_hash = _sha256_file(source_path)

    report = run_validation(source_path=source_path, output_dir=ARTIFACT_DIR)
    after_hash = _sha256_file(source_path)

    assert after_hash == before_hash
    assert report["source_sha256_before"] == before_hash
    assert report["source_sha256_after"] == before_hash
    assert (ARTIFACT_DIR / "source-hash-before.txt").read_text().strip() == before_hash
    assert (ARTIFACT_DIR / "source-hash-after.txt").read_text().strip() == before_hash

    report_path = ARTIFACT_DIR / "differential-report.json"
    report_from_disk = json.loads(report_path.read_text(encoding="utf-8"))
    assert report_from_disk == report

    first_report_bytes = report_path.read_bytes()
    second_report = run_validation(source_path=source_path, output_dir=ARTIFACT_DIR)
    assert second_report == report
    assert report_path.read_bytes() == first_report_bytes
    assert _sha256_file(source_path) == before_hash

    cases = {case["case_id"]: case for case in report["cases"]}
    assert set(cases) == {
        "hist-2024-01-row2-match",
        "hist-2022-11-row3-rounding-reference",
    }
    assert cases["hist-2024-01-row2-match"]["reason_code"] == "match"
    assert cases["hist-2024-01-row2-match"]["canonical_result"]["accrued_interest"] == "10763.89"
    assert cases["hist-2024-01-row2-match"]["historical_value"]["value"] == "10763.89"
    assert cases["hist-2024-01-row2-match"]["delta"] == "0.00"
    assert cases["hist-2022-11-row3-rounding-reference"]["reason_code"] == "rounding_difference"
    assert cases["hist-2022-11-row3-rounding-reference"]["canonical_result"]["accrued_interest"] == "7291.67"
    assert cases["hist-2022-11-row3-rounding-reference"]["historical_value"]["value"] == "7291.66666666667"
    assert cases["hist-2022-11-row3-rounding-reference"]["delta"] == "0.00333333333"

    for case in report["cases"]:
        assert case["reason_code"] in ALLOWED_REASON_CODES
        assert case["source_evidence"]
        assert all(item["sheet"] and item["cell"] for item in case["source_evidence"])
        assert case["historical_value"]["sheet"]
        assert case["historical_value"]["cell"]


def test_manifest_is_explicit_without_dynamic_layout_inference() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))

    assert manifest["version"] == 1
    assert len(manifest["cases"]) == 2
    manifest_text = json.dumps(manifest, ensure_ascii=False).lower()
    for forbidden in ("alias", "infer", "discover", "header_map", "layout_map"):
        assert forbidden not in manifest_text

    for case in manifest["cases"]:
        assert case["case_id"]
        assert case["calculation_month"]
        assert case["source_evidence"]
        assert case["historical_value_cell"]["sheet"]
        assert case["historical_value_cell"]["cell"]
        assert case["canonical_input"]["loan"]["loan_id"]
        assert case["canonical_input"]["loan"]["opening_principal"]
        assert case["canonical_input"]["loan"]["annual_rate"]
        assert case["canonical_input"]["loan"]["day_count_basis"] in (360, 365)
        assert case["expected_reason_code"] in ALLOWED_REASON_CODES


def test_production_importer_accepts_known_historical_month_read_only() -> None:
    source_path = _historical_workbook_path()
    before_hash = _sha256_file(source_path)

    result = import_workbook(
        filename=source_path.name,
        source_bytes=source_path.read_bytes(),
        period=NaturalMonth(2024, 1),
    )

    assert _sha256_file(source_path) == before_hash
    assert result.source_sha256 == before_hash
    assert result.errors == ()
    assert result.calculable_input is not None
    rows = {loan.loan_id: loan for loan in result.calculable_input.loans}
    assert "历史:24全年:2" in rows
    assert rows["历史:24全年:2"].company_name == "上海翰凌医疗器械有限公司"
    assert rows["历史:24全年:2"].opening_principal == Decimal("5000000")
    assert rows["历史:24全年:2"].annual_rate == Decimal("0.025")


def test_historical_upload_calculates_and_exports_expected_reference_value() -> None:
    source_path = _historical_workbook_path()
    payload = source_path.read_bytes()
    before_hash = _sha256_file(source_path)

    calculated = calculate_submission(source_path.name, payload, NaturalMonth(2024, 1))
    exported = export_submission(source_path.name, payload, NaturalMonth(2024, 1))

    assert _sha256_file(source_path) == before_hash
    assert calculated.errors == ()
    assert calculated.calculation is not None
    rows = {row.loan_id: row for row in calculated.calculation.loan_rows}
    assert rows["历史:24全年:2"].accrued_interest == Decimal("10763.89")
    assert exported.errors == ()
    assert exported.output is not None
    workbook = load_workbook(BytesIO(exported.output.workbook_bytes), data_only=False)
    try:
        sheet = workbook[RESULT_SHEET]
        formulas = [
            cell.coordinate
            for row in sheet.iter_rows()
            for cell in row
            if type(cell.value) is str and cell.value.startswith("=")
        ]
        assert formulas == []
        headers = [cell.value for cell in sheet[1]]
        loan_id_col = headers.index("贷款ID") + 1
        interest_col = headers.index("当月计提利息（元）") + 1
        exported_rows = {
            sheet.cell(row, loan_id_col).value: sheet.cell(row, interest_col).value
            for row in range(2, sheet.max_row + 1)
        }
        assert Decimal(str(exported_rows["历史:24全年:2"])) == Decimal("10763.89")
    finally:
        workbook.close()


def test_historical_workbook_without_selected_month_does_not_report_template_sheets() -> None:
    source_path = _historical_workbook_path()
    before_hash = _sha256_file(source_path)

    result = import_workbook(
        filename=source_path.name,
        source_bytes=source_path.read_bytes(),
        period=NaturalMonth(2026, 7),
    )

    assert _sha256_file(source_path) == before_hash
    assert result.calculable_input is None
    assert [error.error_code for error in result.errors] == [
        WorkbookErrorCode.HISTORICAL_PERIOD_NOT_FOUND
    ]
