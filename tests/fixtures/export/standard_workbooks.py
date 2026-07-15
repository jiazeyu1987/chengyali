from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook

from loan_interest_accrual.workbook import (
    LOAN_SHEET,
    LOAN_TEMPLATE_HEADERS,
    MOVEMENT_SHEET,
    MOVEMENT_TEMPLATE_HEADERS,
)


def loan_row(
    loan_id: object = "L-001",
    company: object = "甲公司",
    contract: object = "HT-001",
    bank: object = "甲银行",
    principal: object = 1000,
    rate: object = 0.025,
    basis: object = 365,
    start: object = date(2025, 1, 1),
    end: object = date(2025, 12, 31),
    capitalized: object = "否",
    note: object = None,
) -> list[object]:
    return [
        company,
        contract,
        bank,
        principal,
        rate,
        start,
        note,
    ]


def movement_row(
    loan_id: object = "L-001",
    event_date: object = date(2025, 6, 15),
    movement_type: object = "放款",
    amount: object = 1,
    note: object = None,
) -> list[object]:
    return [loan_id, event_date, movement_type, amount, note]


def workbook_bytes(
    *,
    loans: list[list[object]] | None = None,
    movements: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    loan_sheet = workbook.create_sheet(LOAN_SHEET)
    loan_sheet.append(list(LOAN_TEMPLATE_HEADERS))
    for values in loans if loans is not None else [loan_row()]:
        loan_sheet.append(values)

    rate_column = LOAN_TEMPLATE_HEADERS.index("年利率") + 1
    date_columns = [LOAN_TEMPLATE_HEADERS.index("借款时间") + 1]
    for row in range(2, loan_sheet.max_row + 1):
        loan_sheet.cell(row, rate_column).number_format = "0.00%"
        for column in date_columns:
            loan_sheet.cell(row, column).number_format = "yyyy-mm-dd"

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
