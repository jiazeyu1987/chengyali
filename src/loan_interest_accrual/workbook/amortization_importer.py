from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256

from openpyxl.utils.datetime import from_excel

from loan_interest_accrual.domain import AmortizationAsset, NaturalMonth

from .amortization_schema import (
    AMORTIZATION_HEADERS,
    AMORTIZATION_SHEET,
    MAX_AMORTIZATION_ROWS,
    AmortizationWorkbookImportResult,
    AmortizationWorkbookInput,
)
from .safe_reader import read_workbook
from .schema import WorkbookError, WorkbookErrorCode


def _error(
    code: WorkbookErrorCode,
    row: int | None,
    field: str,
    message: str,
) -> WorkbookError:
    return WorkbookError(code, AMORTIZATION_SHEET, row, field, message)


def _text(value: object) -> str | None:
    if type(value) is not str:
        return None
    normalized = value.strip()
    return normalized or None


def _decimal(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        parsed = Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed.is_finite() else None


def _optional_residual(value: object) -> Decimal | None:
    if value is None:
        return Decimal("0")
    if type(value) is str and value.strip() in {"", "-", "—"}:
        return Decimal("0")
    return _decimal(value)


def _date(value: object, epoch) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if type(value) is date:
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch)
        except (OverflowError, TypeError, ValueError):
            return None
        return converted.date() if isinstance(converted, datetime) else converted
    if type(value) is str:
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _positive_int(value: object) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    if not parsed.is_finite() or parsed != parsed.to_integral_value() or parsed <= 0:
        return None
    return int(parsed)


def import_amortization_workbook(
    filename: str,
    source_bytes: bytes,
    period: NaturalMonth,
) -> AmortizationWorkbookImportResult:
    source_hash = sha256(source_bytes).hexdigest()
    safe = read_workbook(filename, source_bytes)
    if safe.workbook is None:
        return AmortizationWorkbookImportResult(
            source_bytes, source_hash, None, safe.errors
        )

    workbook = safe.workbook
    errors = list(safe.errors)
    assets: list[AmortizationAsset] = []
    try:
        if AMORTIZATION_SHEET not in workbook.sheetnames:
            errors.append(
                _error(
                    WorkbookErrorCode.SHEET_MISSING,
                    None,
                    AMORTIZATION_SHEET,
                    f"missing required worksheet: {AMORTIZATION_SHEET}",
                )
            )
        else:
            sheet = workbook[AMORTIZATION_SHEET]
            headers = [cell.value for cell in sheet[1]]
            text_headers = [value.strip() if type(value) is str else value for value in headers]
            counts = Counter(value for value in text_headers if value is not None)
            for header in AMORTIZATION_HEADERS:
                if counts[header] == 0:
                    errors.append(
                        _error(
                            WorkbookErrorCode.COLUMN_MISSING,
                            1,
                            header,
                            f"missing required column: {header}",
                        )
                    )
                elif counts[header] > 1:
                    errors.append(
                        _error(
                            WorkbookErrorCode.COLUMN_DUPLICATE,
                            1,
                            header,
                            f"duplicated column: {header}",
                        )
                    )
            if not errors:
                header_positions = {
                    header: column
                    for column, header in enumerate(text_headers, start=1)
                    if header in AMORTIZATION_HEADERS
                }
                data_row_count = 0
                for row in range(2, sheet.max_row + 1):
                    cells = {
                        header: sheet.cell(row, header_positions[header])
                        for header in AMORTIZATION_HEADERS
                    }
                    if not any(cell.value is not None for cell in cells.values()):
                        continue
                    data_row_count += 1
                    if data_row_count > MAX_AMORTIZATION_ROWS:
                        continue

                    values = {header: cell.value for header, cell in cells.items()}
                    row_errors: list[WorkbookError] = []
                    for header, cell in cells.items():
                        if cell.data_type == "f":
                            row_errors.append(
                                _error(
                                    WorkbookErrorCode.FORMULA_NOT_ALLOWED,
                                    row,
                                    header,
                                    "input cells must contain fixed values",
                                )
                            )
                    primary_category = _text(values["一级分类"])
                    name = _text(values["名称"])
                    expense = _text(values["对应费用"])
                    for header, parsed in (
                        ("一级分类", primary_category),
                        ("名称", name),
                        ("对应费用", expense),
                    ):
                        if parsed is None:
                            row_errors.append(
                                _error(
                                    WorkbookErrorCode.REQUIRED_VALUE_MISSING,
                                    row,
                                    header,
                                    f"{header} must be non-empty text",
                                )
                            )
                    original = _decimal(values["原值"])
                    residual = _optional_residual(values["残值"])
                    if original is None:
                        row_errors.append(
                            _error(
                                WorkbookErrorCode.ORIGINAL_VALUE_INVALID,
                                row,
                                "原值",
                                "original value must be a valid number",
                            )
                        )
                    residual_invalid = residual is None
                    if original is not None and residual is not None:
                        residual_invalid = (
                            original >= 0 and not (0 <= residual <= original)
                        ) or (
                            original < 0 and not (original <= residual <= 0)
                        )
                    if residual_invalid:
                        row_errors.append(
                            _error(
                                WorkbookErrorCode.RESIDUAL_VALUE_INVALID,
                                row,
                                "残值",
                                "residual value must follow the original value's sign direction",
                            )
                        )
                    start = _date(values["开始摊销月"], workbook.epoch)
                    booking = _date(values["入账月"], workbook.epoch)
                    if start is None:
                        row_errors.append(
                            _error(WorkbookErrorCode.DATE_INVALID, row, "开始摊销月", "invalid amortization start date")
                        )
                    elif (start.year, start.month) > (period.year, period.month):
                        row_errors.append(
                            _error(
                                WorkbookErrorCode.START_MONTH_AFTER_CALCULATION,
                                row,
                                "开始摊销月",
                                "amortization start month is after calculation month",
                            )
                        )
                    if booking is None:
                        row_errors.append(
                            _error(WorkbookErrorCode.DATE_INVALID, row, "入账月", "invalid booking month")
                        )
                    term = _positive_int(values["摊销期限/月"])
                    if term is None:
                        row_errors.append(
                            _error(
                                WorkbookErrorCode.AMORTIZATION_TERM_INVALID,
                                row,
                                "摊销期限/月",
                                "amortization term must be a positive integer",
                            )
                        )
                    errors.extend(row_errors)
                    if not row_errors:
                        assets.append(
                            AmortizationAsset(
                                primary_category=primary_category,
                                name=name,
                                expense_category=expense,
                                original_value=original,
                                residual_value=residual,
                                amortization_start=start,
                                booking_month=booking,
                                amortization_term_months=term,
                            )
                        )

                if data_row_count > MAX_AMORTIZATION_ROWS:
                    errors.append(
                        _error(
                            WorkbookErrorCode.ASSET_ROW_LIMIT_EXCEEDED,
                            None,
                            AMORTIZATION_SHEET,
                            f"asset rows exceed {MAX_AMORTIZATION_ROWS}",
                        )
                    )
                if data_row_count == 0:
                    errors.append(
                        _error(
                            WorkbookErrorCode.REQUIRED_VALUE_MISSING,
                            None,
                            AMORTIZATION_SHEET,
                            "at least one asset row is required",
                        )
                    )
    finally:
        workbook.close()

    normalized_errors = tuple(
        sorted(
            errors,
            key=lambda error: (
                error.sheet or "",
                error.row or 0,
                error.column_or_field,
                error.error_code.value,
            ),
        )
    )
    calculable = None
    if not normalized_errors:
        calculable = AmortizationWorkbookInput(period=period, assets=tuple(assets))
    return AmortizationWorkbookImportResult(
        source_bytes=source_bytes,
        source_sha256=source_hash,
        calculable_input=calculable,
        errors=normalized_errors,
    )
