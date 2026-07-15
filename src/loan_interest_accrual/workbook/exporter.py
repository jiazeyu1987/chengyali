from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from loan_interest_accrual.application.results import CalculationResult, calculation_month
from loan_interest_accrual.domain.decimal_math import exact_sum, quantize_money
from loan_interest_accrual.version import __version__

from .export_schema import (
    CAPITALIZATION_SUMMARY_HEADERS,
    CAPITALIZATION_SUMMARY_SHEET,
    CHECK_HEADERS,
    CHECK_SHEET,
    COMPANY_SUMMARY_HEADERS,
    COMPANY_SUMMARY_SHEET,
    EXPORT_SHEET_NAMES,
    PARAMETER_HEADERS,
    PARAMETER_SHEET,
    RESULT_HEADERS,
    RESULT_SHEET,
    SEGMENT_HEADERS,
    SEGMENT_SHEET,
)


class ExportInvariantError(ValueError):
    pass


_NAVY = "18324B"
_BLUE = "1769AA"
_LINE = "D9DEE4"
_PERIOD_FILL = "E8EEF3"
_HEADER_FILL = "F4F6F8"
_COMPANY_RESULT_HEADERS = (
    "公司名称",
    "本金合计（元）",
    "利息合计（元）",
)


def _set_cell(cell: Cell, value: object) -> None:
    cell.value = value
    if type(value) is str:
        cell.data_type = "s"


def _append_row(sheet, values: tuple[object, ...] | list[object]) -> None:
    row_number = sheet.max_row + 1
    for column, value in enumerate(values, start=1):
        _set_cell(sheet.cell(row_number, column), value)


def _write_row_at(
    sheet,
    row_number: int,
    start_column: int,
    values: tuple[object, ...] | list[object],
) -> None:
    for offset, value in enumerate(values):
        _set_cell(sheet.cell(row_number, start_column + offset), value)


def _write_sheet(sheet, headers: tuple[str, ...], rows: list[tuple[object, ...]]) -> None:
    for column, header in enumerate(headers, start=1):
        _set_cell(sheet.cell(1, column), header)
    for row in rows:
        _append_row(sheet, row)
    _format_sheet(sheet, headers)


def _format_sheet(sheet, headers: tuple[str, ...]) -> None:
    for index, header in enumerate(headers, start=1):
        column_letter = sheet.cell(1, index).column_letter
        if "日期" in header or header == "借款时间":
            number_format = "yyyy-mm-dd"
        elif header == "年利率":
            number_format = "0.000000%"
        elif header == "未舍入分段利息（元）":
            number_format = "0.000000000000"
        elif "（元）" in header:
            number_format = '#,##0.00'
        else:
            number_format = None
        if number_format is not None:
            for cell in sheet[column_letter][1:]:
                cell.number_format = number_format
        sheet.column_dimensions[column_letter].width = max(12, min(len(header) + 4, 32))


def _result_rows(calculation: CalculationResult) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = []
    for result in calculation.portfolio_result.loan_results:
        if len(result.segments) <= 1:
            display_rows = (
                (
                    result.loan.opening_principal,
                    result.interest_days,
                    result.accrued_interest,
                ),
            )
        else:
            rounded_interest = [
                quantize_money(segment.unrounded_interest)
                for segment in result.segments
            ]
            rounded_interest[-1] += result.accrued_interest - exact_sum(
                rounded_interest
            )
            display_rows = tuple(
                (segment.principal, segment.days, interest)
                for segment, interest in zip(
                    result.segments, rounded_interest, strict=True
                )
            )
        for principal, days, interest in display_rows:
            rows.append(
                (
                    len(rows) + 1,
                    result.loan.company_name,
                    result.loan.bank_name,
                    principal,
                    result.loan.annual_rate,
                    result.loan.accrual_start,
                    interest,
                    days,
                )
            )
    rows.append(
        (
            "合计",
            "",
            "",
            exact_sum(row.opening_principal for row in calculation.loan_rows),
            "",
            "",
            exact_sum(row.accrued_interest for row in calculation.loan_rows),
            "",
        )
    )
    return rows


def _company_result_rows(
    calculation: CalculationResult,
) -> list[tuple[object, ...]]:
    summary_by_company = {
        row.company_name: row for row in calculation.company_summary_rows
    }
    return [
        (
            company_name,
            summary_by_company[company_name].opening_principal,
            summary_by_company[company_name].accrued_interest,
        )
        for company_name in dict.fromkeys(
            row.company_name for row in calculation.loan_rows
        )
    ]


def _write_result_sheet(sheet, calculation: CalculationResult) -> int:
    period = calculation.period
    period_text = (
        f"计提区间  {period.start_date.isoformat()} 至 "
        f"{period.end_date.isoformat()}"
    )

    sheet.merge_cells("A1:H1")
    _set_cell(sheet["A1"], period_text)
    _write_row_at(sheet, 3, 1, RESULT_HEADERS)

    result_rows = _result_rows(calculation)
    for row_number, values in enumerate(result_rows, start=4):
        _write_row_at(sheet, row_number, 1, values)

    sheet.merge_cells("J1:L1")
    sheet.merge_cells("J2:L2")
    _set_cell(sheet["J1"], "按公司汇总")
    _set_cell(sheet["J2"], "公司本金与利息合计")
    _write_row_at(sheet, 3, 10, _COMPANY_RESULT_HEADERS)
    for row_number, values in enumerate(
        _company_result_rows(calculation), start=4
    ):
        _write_row_at(sheet, row_number, 10, values)

    return 3 + len(result_rows)


def _format_result_sheet(
    sheet,
    calculation: CalculationResult,
    total_row: int,
) -> None:
    thin_line = Side(style="thin", color=_LINE)
    total_line = Side(style="thin", color="64748B")
    period_fill = PatternFill("solid", fgColor=_PERIOD_FILL)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)

    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.freeze_panes = "A4"
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 32

    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
        for cell in row:
            cell.fill = period_fill
            cell.border = Border(bottom=Side(style="thin", color="B8C5D1"))
    sheet["A1"].font = Font(name="Microsoft YaHei", size=15, bold=True, color=_NAVY)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    for row_number in (1, 2):
        for cell in sheet.iter_rows(
            min_row=row_number,
            max_row=row_number,
            min_col=10,
            max_col=12,
        ):
            for item in cell:
                item.fill = header_fill
                item.border = Border(bottom=thin_line)
    sheet["J1"].font = Font(name="Microsoft YaHei", size=10, bold=True, color=_BLUE)
    sheet["J2"].font = Font(name="Microsoft YaHei", size=14, bold=True, color=_NAVY)
    for coordinate in ("J1", "J2"):
        sheet[coordinate].alignment = Alignment(horizontal="left", vertical="center")

    for start_column, end_column in ((1, 8), (10, 12)):
        for cell in sheet.iter_rows(
            min_row=3,
            max_row=3,
            min_col=start_column,
            max_col=end_column,
        ):
            for item in cell:
                item.fill = header_fill
                item.font = Font(name="Microsoft YaHei", size=10, bold=True, color="3F4954")
                item.border = Border(bottom=thin_line)
                item.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    for row in sheet.iter_rows(min_row=4, max_row=total_row, min_col=1, max_col=8):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, color="17202A")
            cell.border = Border(bottom=thin_line)
            cell.alignment = Alignment(vertical="center")
    for cell in sheet[total_row][0:8]:
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="17202A")
        cell.border = Border(top=total_line)

    company_end_row = 3 + len(calculation.company_summary_rows)
    for row in sheet.iter_rows(
        min_row=4,
        max_row=company_end_row,
        min_col=10,
        max_col=12,
    ):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, color="17202A")
            cell.border = Border(bottom=thin_line)
            cell.alignment = Alignment(vertical="center")

    for row_number in range(4, total_row + 1):
        sheet.cell(row_number, 4).number_format = "#,##0.00"
        sheet.cell(row_number, 5).number_format = "0.0000%"
        sheet.cell(row_number, 6).number_format = "yyyy-mm-dd"
        sheet.cell(row_number, 7).number_format = "#,##0.00"
        sheet.cell(row_number, 1).alignment = Alignment(horizontal="center")
        for column in (4, 5, 7, 8):
            sheet.cell(row_number, column).alignment = Alignment(horizontal="right")
    for row_number in range(4, company_end_row + 1):
        for column in (11, 12):
            sheet.cell(row_number, column).number_format = "#,##0.00"
            sheet.cell(row_number, column).alignment = Alignment(horizontal="right")
        sheet.cell(row_number, 10).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

    widths = {
        "A": 8,
        "B": 30,
        "C": 18,
        "D": 18,
        "E": 12,
        "F": 14,
        "G": 20,
        "H": 20,
        "I": 3,
        "J": 30,
        "K": 18,
        "L": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _segment_rows(calculation: CalculationResult) -> list[tuple[object, ...]]:
    return [
        (
            row.calculation_month,
            row.loan_id,
            row.sequence,
            row.start_date,
            row.end_date,
            row.days,
            row.principal,
            row.annual_rate,
            row.day_count_basis,
            row.unrounded_interest,
            row.ending_principal,
            row.trigger_note,
        )
        for row in calculation.segment_rows
    ]


def _company_rows(calculation: CalculationResult) -> list[tuple[object, ...]]:
    return [
        (
            row.company_name,
            row.loan_count,
            row.opening_principal,
            row.total_drawdowns,
            row.total_repayments,
            row.ending_principal,
            row.accrued_interest,
            row.capitalized_interest,
            row.expensed_interest,
        )
        for row in calculation.company_summary_rows
    ]


def _capitalization_rows(calculation: CalculationResult) -> list[tuple[object, ...]]:
    rows = [
        (
            row.company_name,
            row.capitalized_loan_count,
            row.ending_principal,
            row.capitalized_interest,
        )
        for row in calculation.capitalization_summary_rows
    ]
    rows.append(
        (
            "合计",
            sum(row.capitalized_loan_count for row in calculation.capitalization_summary_rows),
            exact_sum(row.ending_principal for row in calculation.capitalization_summary_rows),
            calculation.portfolio_result.total_capitalized_interest,
        )
    )
    return rows


def _check_rows(calculation: CalculationResult) -> list[tuple[object, ...]]:
    return [
        (
            check.display_name,
            check.status,
            check.expected,
            check.actual,
        )
        for check in calculation.checks
    ]


def _parameter_rows(
    calculation: CalculationResult,
    generated_at: datetime,
) -> list[tuple[object, ...]]:
    period = calculation.period
    return [
        (
            "计算区间",
            calculation_month(period),
        ),
        ("期间开始日期", period.start_date.isoformat()),
        ("期间结束日期", period.end_date.isoformat()),
        ("金额单位", "人民币元"),
        ("利率输入格式", "Excel 百分比"),
        ("放款生效规则", "次日起息"),
        ("还本生效规则", "计息至还本当日"),
        ("贷款日期规则", "开始日和结束日均计息"),
        ("舍入规则", "逐笔最终 ROUND_HALF_UP 到 0.01"),
        ("应用版本", __version__),
        ("生成时间", generated_at.isoformat()),
    ]


def _assert_checks_pass(calculation: CalculationResult) -> None:
    failed = tuple(check.display_name for check in calculation.checks if not check.passed)
    if failed:
        raise ExportInvariantError(
            "export requires all checks to pass: " + ", ".join(failed)
        )


def inspect_export_package(workbook_bytes: bytes) -> tuple[str, ...]:
    problems: list[str] = []
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False, keep_links=False)
    try:
        if tuple(workbook.sheetnames) != EXPORT_SHEET_NAMES:
            problems.append("worksheet names do not match export schema")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        problems.append(
                            f"formula cell is present: {sheet.title}!{cell.coordinate}"
                        )
    finally:
        workbook.close()

    with ZipFile(BytesIO(workbook_bytes)) as package:
        for name in package.namelist():
            lowered = name.lower()
            if "vbaproject" in lowered:
                problems.append(f"VBA part is present: {name}")
            if "externallinks" in lowered:
                problems.append(f"external link part is present: {name}")
            if "connections" in lowered:
                problems.append(f"connection part is present: {name}")
            if not (lowered.endswith(".xml") or lowered.endswith(".rels")):
                continue
            content = package.read(name).decode("utf-8", errors="ignore")
            if "<f>" in content:
                problems.append(f"formula element is present: {name}")
            if 'TargetMode="External"' in content:
                problems.append(f"external relationship is present: {name}")
    return tuple(problems)


def export_calculation_workbook(
    calculation: CalculationResult,
    *,
    generated_at: datetime | None = None,
) -> bytes:
    _assert_checks_pass(calculation)
    timestamp = generated_at or datetime.now(timezone.utc).replace(microsecond=0)

    workbook = Workbook()
    workbook.active.title = RESULT_SHEET

    result_sheet = workbook[RESULT_SHEET]
    total_row = _write_result_sheet(result_sheet, calculation)
    _format_result_sheet(result_sheet, calculation, total_row)

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    output = stream.getvalue()
    problems = inspect_export_package(output)
    if problems:
        raise ExportInvariantError("; ".join(problems))
    return output
