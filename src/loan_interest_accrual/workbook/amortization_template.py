from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .amortization_schema import (
    AMORTIZATION_HEADERS,
    AMORTIZATION_SHEET,
    AMOUNT_NUMBER_FORMAT,
    DATE_NUMBER_FORMAT,
    MAX_AMORTIZATION_ROWS,
    MONTH_COUNT_NUMBER_FORMAT,
)


_INSTRUCTIONS = (
    "填写资产的一级分类，例如软件、专利、土地使用权或长期待摊费用类别。",
    "填写资产或项目名称。",
    "填写对应费用，例如管理费用或制造费用。",
    "填写资产原值；冲销或更正记录允许填写负数。",
    "残值可留空，留空按0处理；如填写，绝对值不得超过原值。",
    "填写开始摊销月份中的任一有效日期，计算时按自然月处理。",
    "填写入账月份中的任一有效日期，仅用于结果展示。",
    "填写正整数月份，例如36、60或120。",
)


def generate_amortization_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = AMORTIZATION_SHEET
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:H{MAX_AMORTIZATION_ROWS + 1}"

    yellow = PatternFill("solid", fgColor="FFF200")
    border = Border(
        left=Side(style="thin", color="6B7280"),
        right=Side(style="thin", color="6B7280"),
        top=Side(style="thin", color="6B7280"),
        bottom=Side(style="thin", color="6B7280"),
    )
    for column, (header, instruction) in enumerate(
        zip(AMORTIZATION_HEADERS, _INSTRUCTIONS, strict=True), start=1
    ):
        cell = sheet.cell(1, column, header)
        cell.fill = yellow
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.comment = Comment(f"填写说明：{instruction}", "本地应用")

    widths = (18, 38, 16, 16, 14, 17, 17, 17)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
    sheet.row_dimensions[1].height = 24

    for column in range(1, 9):
        sheet.cell(2, column).border = border
        sheet.cell(2, column).font = Font(name="Microsoft YaHei", size=10, color="0000FF")
    sheet["D2"].number_format = AMOUNT_NUMBER_FORMAT
    sheet["E2"].number_format = AMOUNT_NUMBER_FORMAT
    sheet["F2"].number_format = DATE_NUMBER_FORMAT
    sheet["G2"].number_format = DATE_NUMBER_FORMAT
    sheet["H2"].number_format = MONTH_COUNT_NUMBER_FORMAT

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()
