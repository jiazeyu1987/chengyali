import ast
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl.workbook.workbook import Workbook

from loan_interest_accrual.domain import (
    DayCountBasis,
    DomainValidationError,
    Loan,
    Movement,
    MovementType,
    NaturalMonth,
    calculate_portfolio,
)

from .limits import MAX_LOAN_ROWS, MAX_MOVEMENT_ROWS
from .schema import (
    LOAN_REQUIRED_HEADERS,
    LOAN_SHEET,
    LOAN_TEMPLATE_HEADERS,
    MOVEMENT_REQUIRED_HEADERS,
    MOVEMENT_SHEET,
    MOVEMENT_TEMPLATE_HEADERS,
    OPTIONAL_HEADERS,
    WorkbookError,
    WorkbookErrorCode,
)


@dataclass(frozen=True, slots=True)
class ValidationOutcome:
    loans: tuple[Loan, ...]
    movements: tuple[Movement, ...]
    errors: tuple[WorkbookError, ...]


def _error(
    code: WorkbookErrorCode,
    sheet: str,
    row: int | None,
    field: str,
    message: str,
) -> WorkbookError:
    return WorkbookError(code, sheet, row, field, message)


def _blank(value: object) -> bool:
    return value is None or (type(value) is str and value.strip() == "")


def _text(value: object) -> str | None:
    return value if type(value) is str and value.strip() != "" else None


def _decimal(value: object) -> Decimal | None:
    if type(value) is bool or not isinstance(value, (int, float, Decimal)):
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return result if result.is_finite() else None


def _safe_principal_formula(value: object) -> Decimal | None:
    if type(value) is not str or not value.startswith("=") or len(value) > 200:
        return None
    try:
        tree = ast.parse(value[1:], mode="eval")
    except (SyntaxError, ValueError):
        return None
    if sum(1 for _ in ast.walk(tree)) > 50:
        return None

    def evaluate(node: ast.AST) -> Decimal:
        if type(node) is ast.Expression:
            return evaluate(node.body)
        if type(node) is ast.Constant and type(node.value) in (int, float):
            return Decimal(str(node.value))
        if type(node) is ast.UnaryOp and type(node.op) in (ast.UAdd, ast.USub):
            operand = evaluate(node.operand)
            return operand if type(node.op) is ast.UAdd else -operand
        if type(node) is ast.BinOp and type(node.op) in (
            ast.Add,
            ast.Sub,
            ast.Mult,
            ast.Div,
        ):
            left = evaluate(node.left)
            right = evaluate(node.right)
            if type(node.op) is ast.Add:
                return left + right
            if type(node.op) is ast.Sub:
                return left - right
            if type(node.op) is ast.Mult:
                return left * right
            return left / right
        raise ValueError("unsupported principal formula")

    try:
        result = evaluate(tree)
    except (ArithmeticError, InvalidOperation, ValueError):
        return None
    return result if result.is_finite() else None


def _date(value: object) -> date | None:
    if type(value) is datetime:
        return value.date()
    return value if type(value) is date else None


_REPAYMENT_ACTION = r"(?:归还|偿还|还|换)\s*本金"
_REPAYMENT_SEPARATOR = r"[，,、:：\s]*"
_NUMERIC_REPAYMENT_DATE = re.compile(
    rf"(?P<year>\d{{2}}|\d{{4}})\s*[./-]\s*(?P<month>\d{{1,2}})\s*"
    rf"[./-]\s*(?P<day>\d{{1,2}})\s*[日号]?{_REPAYMENT_SEPARATOR}"
    rf"{_REPAYMENT_ACTION}"
)
_FULL_REPAYMENT_DATE = re.compile(
    rf"(?P<year>\d{{4}})\s*年\s*(?P<month>\d{{1,2}})\s*月\s*"
    rf"(?P<day>\d{{1,2}})\s*[日号]{_REPAYMENT_SEPARATOR}{_REPAYMENT_ACTION}"
)
_MONTH_DAY_REPAYMENT_DATE = re.compile(
    rf"(?P<month>\d{{1,2}})\s*月\s*(?P<day>\d{{1,2}})\s*[日号]"
    rf"{_REPAYMENT_SEPARATOR}{_REPAYMENT_ACTION}"
)
_DAY_REPAYMENT_DATE = re.compile(
    rf"(?P<day>\d{{1,2}})\s*[日号]{_REPAYMENT_SEPARATOR}{_REPAYMENT_ACTION}"
)
_REPAYMENT_DATE_AFTER_ACTION = re.compile(
    rf"{_REPAYMENT_ACTION}[^\d]{{0,8}}(?P<day>\d{{1,2}})\s*[日号]"
)

_PARTIAL_REPAYMENT_ACTION = r"(?:已\s*)?(?:归还|偿还|还|换)(?:款|本金)?"
_PARTIAL_REPAYMENT_AMOUNT = (
    r"(?P<amount>\d[\d,]*(?:\.\d+)?)\s*"
    r"(?P<unit>亿元|万元|亿|万|元)?(?![\d,.])(?!\s*%)"
)
_NUMERIC_PARTIAL_REPAYMENT = re.compile(
    rf"(?P<year>\d{{2}}|\d{{4}})\s*[./-]\s*(?P<month>\d{{1,2}})\s*"
    rf"[./-]\s*(?P<day>\d{{1,2}})\s*[日号]?{_REPAYMENT_SEPARATOR}"
    rf"{_PARTIAL_REPAYMENT_ACTION}{_REPAYMENT_SEPARATOR}{_PARTIAL_REPAYMENT_AMOUNT}"
)
_FULL_PARTIAL_REPAYMENT = re.compile(
    rf"(?P<year>\d{{4}})\s*年\s*(?P<month>\d{{1,2}})\s*月\s*"
    rf"(?P<day>\d{{1,2}})\s*[日号]{_REPAYMENT_SEPARATOR}"
    rf"{_PARTIAL_REPAYMENT_ACTION}{_REPAYMENT_SEPARATOR}{_PARTIAL_REPAYMENT_AMOUNT}"
)
_MONTH_DAY_PARTIAL_REPAYMENT = re.compile(
    rf"(?P<month>\d{{1,2}})\s*月\s*(?P<day>\d{{1,2}})\s*[日号]"
    rf"{_REPAYMENT_SEPARATOR}{_PARTIAL_REPAYMENT_ACTION}"
    rf"{_REPAYMENT_SEPARATOR}{_PARTIAL_REPAYMENT_AMOUNT}"
)
_DAY_PARTIAL_REPAYMENT = re.compile(
    rf"(?P<day>\d{{1,2}})\s*[日号]{_REPAYMENT_SEPARATOR}"
    rf"{_PARTIAL_REPAYMENT_ACTION}{_REPAYMENT_SEPARATOR}{_PARTIAL_REPAYMENT_AMOUNT}"
)


def _partial_repayment_from_note(
    value: object,
    period: NaturalMonth,
) -> tuple[date, Decimal] | None:
    if type(value) is not str:
        return None
    note = value.strip()
    if not note:
        return None

    match = None
    repayment_date = None
    for pattern, date_kind in (
        (_NUMERIC_PARTIAL_REPAYMENT, "numeric"),
        (_FULL_PARTIAL_REPAYMENT, "full"),
        (_MONTH_DAY_PARTIAL_REPAYMENT, "month_day"),
        (_DAY_PARTIAL_REPAYMENT, "day"),
    ):
        match = pattern.search(note)
        if match is None:
            continue
        year = period.end_date.year
        month = period.end_date.month
        if date_kind in {"numeric", "full"}:
            year = int(match.group("year"))
            if year < 100:
                year += 2000
            month = int(match.group("month"))
        elif date_kind == "month_day":
            month = int(match.group("month"))
        repayment_date = date(year, month, int(match.group("day")))
        break

    if match is None or repayment_date is None:
        return None

    amount = Decimal(match.group("amount").replace(",", ""))
    multiplier = {
        "亿元": Decimal("100000000"),
        "亿": Decimal("100000000"),
        "万元": Decimal("10000"),
        "万": Decimal("10000"),
        "元": Decimal("1"),
        None: Decimal("1"),
    }[match.group("unit")]
    return repayment_date, amount * multiplier


def _repayment_date_from_note(
    value: object,
    period: NaturalMonth,
) -> date | None:
    if type(value) is not str:
        return None
    note = value.strip()
    if not note:
        return None

    match = _NUMERIC_REPAYMENT_DATE.search(note)
    if match is not None:
        year = int(match.group("year"))
        if year < 100:
            year += 2000
        return date(year, int(match.group("month")), int(match.group("day")))

    match = _FULL_REPAYMENT_DATE.search(note)
    if match is not None:
        return date(
            int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        )

    match = _MONTH_DAY_REPAYMENT_DATE.search(note)
    if match is not None:
        return date(
            period.end_date.year,
            int(match.group("month")),
            int(match.group("day")),
        )

    match = _DAY_REPAYMENT_DATE.search(note)
    if match is None:
        match = _REPAYMENT_DATE_AFTER_ACTION.search(note)
    if match is None:
        return None
    return date(
        period.end_date.year,
        period.end_date.month,
        int(match.group("day")),
    )


def _headers(
    sheet,
    required: tuple[str, ...],
    optional: tuple[str, ...] = (),
) -> tuple[dict[str, int], list[WorkbookError]]:
    values = [cell.value for cell in sheet[1]]
    positions: dict[str, list[int]] = {}
    for index, value in enumerate(values, start=1):
        if type(value) is str:
            positions.setdefault(value, []).append(index)
    errors: list[WorkbookError] = []
    mapping: dict[str, int] = {}
    for header in required + optional:
        found = positions.get(header, [])
        if not found:
            if header in required:
                errors.append(
                    _error(
                        WorkbookErrorCode.COLUMN_MISSING,
                        sheet.title,
                        1,
                        header,
                        f"required column is missing: {header}",
                    )
                )
        elif len(found) > 1:
            errors.append(
                _error(
                    WorkbookErrorCode.COLUMN_DUPLICATE,
                    sheet.title,
                    1,
                    header,
                    f"required column is duplicated: {header}",
                )
            )
        else:
            mapping[header] = found[0]
    return mapping, errors


def _data_row_numbers(sheet) -> tuple[int, ...]:
    return tuple(
        sorted(
            {
                cell.row
                for cell in sheet._cells.values()
                if cell.row >= 2 and not _blank(cell.value)
            }
        )
    )


def _formula_errors(sheet) -> list[WorkbookError]:
    errors: list[WorkbookError] = []
    headers = {
        index: value
        for index, value in enumerate(
            (cell.value for cell in sheet[1]), start=1
        )
        if type(value) is str
    }
    for cell in sorted(sheet._cells.values(), key=lambda item: (item.row, item.column)):
        if cell.data_type == "f":
            field = headers.get(cell.column, cell.coordinate)
            if field == "期初本金（元）" and _safe_principal_formula(cell.value) is not None:
                continue
            errors.append(
                _error(
                    WorkbookErrorCode.FORMULA_NOT_ALLOWED,
                    sheet.title,
                    cell.row,
                    field,
                    "formula cells are not allowed",
                )
            )
    return errors


def _loan_rows(
    sheet,
    period: NaturalMonth,
    mapping: dict[str, int],
    formula_errors: tuple[WorkbookError, ...],
):
    errors: list[WorkbookError] = []
    parsed: list[tuple[int, dict[str, object]]] = []
    formula_cells = {
        (error.row, error.column_or_field) for error in formula_errors
    }
    for row in _data_row_numbers(sheet):
        values = {
            field: sheet.cell(row, column).value for field, column in mapping.items()
        }
        valid = True
        loan_id = f"ROW-{row:06d}"

        start_value = values.get("借款时间")
        start = _date(start_value)
        # Rows that have not started yet, or whose remarks show full principal
        # repayment before the selected range, are simply outside this range's
        # result set. They are skipped instead of blocking all valid rows.
        if start is not None and start > period.end_date:
            continue
        try:
            partial_repayment = _partial_repayment_from_note(
                values.get("备注"), period
            )
            repayment_end = (
                None
                if partial_repayment is not None
                else _repayment_date_from_note(values.get("备注"), period)
            )
        except ValueError:
            errors.append(
                _error(
                    WorkbookErrorCode.DATE_INVALID,
                    LOAN_SHEET,
                    row,
                    "备注",
                    "备注中的归还本金日期无效",
                )
            )
            repayment_end = None
            partial_repayment = None
            valid = False
        if repayment_end is not None and repayment_end < period.start_date:
            continue

        for field in ("公司名称",):
            if (row, field) in formula_cells:
                valid = False
            elif _text(values.get(field)) is None:
                errors.append(
                    _error(
                        WorkbookErrorCode.REQUIRED_VALUE_MISSING,
                        LOAN_SHEET,
                        row,
                        field,
                        f"{field} must be non-empty text",
                    )
                )
                valid = False

        principal_value = values.get("期初本金（元）")
        principal = (
            _safe_principal_formula(principal_value)
            if type(principal_value) is str and principal_value.startswith("=")
            else _decimal(principal_value)
        )
        if (row, "期初本金（元）") in formula_cells:
            valid = False
        elif principal is None:
            errors.append(
                _error(
                    WorkbookErrorCode.VALUE_TYPE_INVALID,
                    LOAN_SHEET,
                    row,
                    "期初本金（元）",
                    "opening principal must be a numeric RMB amount",
                )
            )
            valid = False
        elif principal < 0:
            errors.append(
                _error(
                    WorkbookErrorCode.NEGATIVE_PRINCIPAL,
                    LOAN_SHEET,
                    row,
                    "期初本金（元）",
                    "opening principal must not be negative",
                )
            )
            valid = False

        rate_cell = sheet.cell(row, mapping.get("年利率", 1))
        rate = _decimal(values.get("年利率"))
        if (row, "年利率") in formula_cells:
            valid = False
        elif (
            rate is None
            or rate <= 0
            or rate >= 1
            or "%" not in rate_cell.number_format
        ):
            errors.append(
                _error(
                    WorkbookErrorCode.INTEREST_RATE_INVALID,
                    LOAN_SHEET,
                    row,
                    "年利率",
                    "annual rate must be an Excel percentage value greater than 0 and less than 100%",
                )
            )
            valid = False

        basis = DayCountBasis.DAYS_360

        # 除备注明确写明区间内归还本金日期外，计提截止日固定为
        # 所选日期区间的结束日。
        end = period.end_date
        if repayment_end is not None and period.contains(repayment_end):
            end = repayment_end
        if (row, "借款时间") in formula_cells:
            valid = False

        partial_repayment_date = None
        partial_repayment_amount = None
        if partial_repayment is not None:
            partial_repayment_date, partial_repayment_amount = partial_repayment
            if partial_repayment_amount <= 0:
                errors.append(
                    _error(
                        WorkbookErrorCode.MOVEMENT_AMOUNT_INVALID,
                        LOAN_SHEET,
                        row,
                        "备注",
                        "备注中的还本金额必须大于0",
                    )
                )
                valid = False
            elif principal is not None and partial_repayment_amount > principal:
                errors.append(
                    _error(
                        WorkbookErrorCode.NEGATIVE_PRINCIPAL,
                        LOAN_SHEET,
                        row,
                        "备注",
                        "备注中的还本金额不得大于借款本金",
                    )
                )
                valid = False
        elif _blank(start_value):
            errors.append(
                _error(
                    WorkbookErrorCode.REQUIRED_VALUE_MISSING,
                    LOAN_SHEET,
                    row,
                    "借款时间",
                    "借款时间 is required",
                )
            )
            valid = False
        elif start is None:
            errors.append(
                _error(
                    WorkbookErrorCode.DATE_INVALID,
                    LOAN_SHEET,
                    row,
                    "借款时间",
                    "借款时间 must be a valid Excel date",
                )
            )
            valid = False
        if start is not None and end is not None:
            if repayment_end is not None and end < start:
                errors.append(
                    _error(
                        WorkbookErrorCode.DATE_RANGE_INVALID,
                        LOAN_SHEET,
                        row,
                        "备注",
                        "备注中的归还本金日期不得早于借款时间",
                    )
                )
                valid = False
            if (
                partial_repayment_date is not None
                and partial_repayment_date < start
            ):
                errors.append(
                    _error(
                        WorkbookErrorCode.DATE_RANGE_INVALID,
                        LOAN_SHEET,
                        row,
                        "备注",
                        "备注中的还本日期不得早于借款时间",
                    )
                )
                valid = False

        adjusted_principal = principal
        if (
            valid
            and partial_repayment_date is not None
            and partial_repayment_amount is not None
            and partial_repayment_date < period.start_date
        ):
            adjusted_principal = principal - partial_repayment_amount

        parsed.append(
            (
                row,
                {
                    "valid": valid,
                    "loan_id": loan_id,
                    "company_name": _text(values.get("公司名称")) or "",
                    "contract_number": _text(values.get("贷款合同号")) or "",
                    "bank_name": _text(values.get("贷款银行")) or "",
                    "opening_principal": adjusted_principal,
                    "annual_rate": rate,
                    "day_count_basis": basis,
                    "accrual_start": start,
                    "accrual_end": end,
                    "capitalize_interest": False,
                    "partial_repayment_date": (
                        partial_repayment_date
                        if partial_repayment_date is not None
                        and period.contains(partial_repayment_date)
                        else None
                    ),
                    "partial_repayment_amount": partial_repayment_amount,
                },
            )
        )
    counts = Counter(values["loan_id"] for _, values in parsed)
    duplicate_ids: set[str] = set()
    loans: list[tuple[int, Loan]] = []
    note_movements: list[tuple[int, Movement]] = []
    for row, values in parsed:
        if not values["valid"] or values["loan_id"] in duplicate_ids:
            continue
        try:
            loan = Loan(
                loan_id=values["loan_id"],
                company_name=values["company_name"],
                contract_number=values["contract_number"],
                bank_name=values["bank_name"],
                opening_principal=values["opening_principal"],
                annual_rate=values["annual_rate"],
                day_count_basis=values["day_count_basis"],
                accrual_start=values["accrual_start"],
                accrual_end=values["accrual_end"],
                capitalize_interest=values["capitalize_interest"],
            )
            loans.append((row, loan))
            if values["partial_repayment_date"] is not None:
                note_movements.append(
                    (
                        row,
                        Movement(
                            loan_id=values["loan_id"],
                            event_date=values["partial_repayment_date"],
                            movement_type=MovementType.REPAYMENT,
                            amount=values["partial_repayment_amount"],
                        ),
                    )
                )
        except DomainValidationError:
            continue
    return loans, counts, note_movements, errors


def _movement_rows(
    sheet,
    period: NaturalMonth,
    mapping,
    loan_counts,
    formula_errors: tuple[WorkbookError, ...],
):
    errors: list[WorkbookError] = []
    movements: list[tuple[int, Movement]] = []
    invalid_loan_ids: set[str] = set()
    formula_cells = {
        (error.row, error.column_or_field) for error in formula_errors
    }
    for row in _data_row_numbers(sheet):
        values = {
            field: sheet.cell(row, column).value for field, column in mapping.items()
        }
        valid = True
        loan_id = _text(values.get("贷款ID"))
        if (row, "贷款ID") in formula_cells:
            valid = False
        elif loan_id is None:
            errors.append(
                _error(
                    WorkbookErrorCode.MOVEMENT_LOAN_ID_REQUIRED,
                    MOVEMENT_SHEET,
                    row,
                    "贷款ID",
                    "movement 贷款ID must be non-empty text",
                )
            )
            valid = False
        elif loan_counts.get(loan_id, 0) == 0:
            errors.append(
                _error(
                    WorkbookErrorCode.MOVEMENT_LOAN_ID_NOT_FOUND,
                    MOVEMENT_SHEET,
                    row,
                    "贷款ID",
                    f"movement 贷款ID was not found: {loan_id}",
                )
            )
            valid = False
        elif loan_counts[loan_id] > 1:
            errors.append(
                _error(
                    WorkbookErrorCode.MOVEMENT_LOAN_ID_AMBIGUOUS,
                    MOVEMENT_SHEET,
                    row,
                    "贷款ID",
                    f"movement 贷款ID is not unique: {loan_id}",
                )
            )
            valid = False

        event_date = _date(values.get("变动日期"))
        if (row, "变动日期") in formula_cells:
            valid = False
        elif event_date is None:
            errors.append(
                _error(
                    WorkbookErrorCode.DATE_INVALID,
                    MOVEMENT_SHEET,
                    row,
                    "变动日期",
                    "movement date must be a valid Excel date",
                )
            )
            valid = False
        elif not period.contains(event_date):
            errors.append(
                _error(
                    WorkbookErrorCode.MOVEMENT_DATE_OUTSIDE_MONTH,
                    MOVEMENT_SHEET,
                    row,
                    "变动日期",
                    "movement date must be inside the selected month",
                )
            )
            valid = False

        type_value = values.get("变动类型")
        movement_type = {
            "放款": MovementType.DRAWDOWN,
            "还本": MovementType.REPAYMENT,
        }.get(type_value)
        if (row, "变动类型") in formula_cells:
            valid = False
        elif movement_type is None:
            errors.append(
                _error(
                    WorkbookErrorCode.MOVEMENT_TYPE_INVALID,
                    MOVEMENT_SHEET,
                    row,
                    "变动类型",
                    "movement type must be 放款 or 还本",
                )
            )
            valid = False

        amount = _decimal(values.get("变动金额（元）"))
        if (row, "变动金额（元）") in formula_cells:
            valid = False
        elif amount is None or amount <= 0:
            errors.append(
                _error(
                    WorkbookErrorCode.MOVEMENT_AMOUNT_INVALID,
                    MOVEMENT_SHEET,
                    row,
                    "变动金额（元）",
                    "movement amount must be a numeric RMB amount greater than zero",
                )
            )
            valid = False
        if not valid:
            if loan_id is not None and loan_counts.get(loan_id) == 1:
                invalid_loan_ids.add(loan_id)
            continue
        movements.append(
            (
                row,
                Movement(
                    loan_id=loan_id,
                    event_date=event_date,
                    movement_type=movement_type,
                    amount=amount,
                ),
            )
        )
    return movements, invalid_loan_ids, errors


def _sort_key(error: WorkbookError):
    sheet_rank = {None: -1, LOAN_SHEET: 0, MOVEMENT_SHEET: 1}
    field_orders = {
        LOAN_SHEET: {name: index for index, name in enumerate(LOAN_TEMPLATE_HEADERS)},
        MOVEMENT_SHEET: {
            name: index for index, name in enumerate(MOVEMENT_TEMPLATE_HEADERS)
        },
    }
    return (
        sheet_rank.get(error.sheet, 2),
        error.row if error.row is not None else 0,
        field_orders.get(error.sheet, {}).get(error.column_or_field, 999),
        error.column_or_field,
        error.error_code.value,
    )


def sort_errors(errors) -> tuple[WorkbookError, ...]:
    return tuple(sorted(errors, key=_sort_key))


def _business_validation_errors(
    period: NaturalMonth,
    loans_with_rows: tuple[tuple[int, Loan], ...],
    movements_with_rows: tuple[tuple[int, Movement], ...],
    invalid_movement_ids: set[str],
) -> tuple[WorkbookError, ...]:
    loans = tuple(
        loan for _, loan in loans_with_rows if loan.loan_id not in invalid_movement_ids
    )
    loan_ids = {loan.loan_id for loan in loans}
    movements = tuple(
        movement
        for _, movement in movements_with_rows
        if movement.loan_id in loan_ids
    )
    if not loans:
        return ()

    loan_rows = {loan.loan_id: row for row, loan in loans_with_rows}
    movement_rows: dict[tuple[str, date], int] = {}
    for row, movement in movements_with_rows:
        movement_rows.setdefault((movement.loan_id, movement.event_date), row)

    try:
        calculate_portfolio(period, loans, movements)
    except DomainValidationError as caught:
        errors: list[WorkbookError] = []
        for domain_error in caught.errors:
            if domain_error.error_code.value == WorkbookErrorCode.NEGATIVE_PRINCIPAL:
                errors.append(
                    _error(
                        WorkbookErrorCode.NEGATIVE_PRINCIPAL,
                        MOVEMENT_SHEET,
                        movement_rows.get(
                            (domain_error.loan_id, domain_error.event_date)
                        ),
                        "变动金额（元）",
                        domain_error.message,
                    )
                )
                continue
            workbook_code = (
                WorkbookErrorCode(domain_error.error_code.value)
                if domain_error.error_code.value in WorkbookErrorCode._value2member_map_
                else WorkbookErrorCode.VALUE_TYPE_INVALID
            )
            errors.append(
                _error(
                    workbook_code,
                    LOAN_SHEET,
                    loan_rows.get(domain_error.loan_id),
                    domain_error.column_or_field,
                    domain_error.message,
                )
            )
        return tuple(errors)
    return ()


def validate_workbook(workbook: Workbook, period: NaturalMonth) -> ValidationOutcome:
    errors: list[WorkbookError] = []
    if LOAN_SHEET not in workbook.sheetnames:
        errors.append(
            _error(
                WorkbookErrorCode.SHEET_MISSING,
                LOAN_SHEET,
                None,
                LOAN_SHEET,
                f"required worksheet is missing: {LOAN_SHEET}",
            )
        )
    if LOAN_SHEET in workbook.sheetnames:
        errors.extend(_formula_errors(workbook[LOAN_SHEET]))
    if LOAN_SHEET not in workbook.sheetnames:
        return ValidationOutcome((), (), sort_errors(errors))

    loan_sheet = workbook[LOAN_SHEET]
    loan_formula_errors = tuple(
        error for error in errors if error.sheet == LOAN_SHEET
    )
    loan_mapping, loan_header_errors = _headers(
        loan_sheet, LOAN_REQUIRED_HEADERS, OPTIONAL_HEADERS
    )
    errors.extend(loan_header_errors)

    loan_over = loan_sheet.max_row - 1 > MAX_LOAN_ROWS
    if loan_over:
        errors.append(
            _error(
                WorkbookErrorCode.LOAN_ROW_LIMIT_EXCEEDED,
                LOAN_SHEET,
                None,
                "rows",
                f"贷款主表 exceeds {MAX_LOAN_ROWS} data rows",
            )
        )
    if loan_header_errors or loan_over:
        return ValidationOutcome((), (), sort_errors(errors))

    loans_with_rows, loan_counts, note_movements, loan_errors = _loan_rows(
        loan_sheet, period, loan_mapping, loan_formula_errors
    )
    errors.extend(loan_errors)
    errors.extend(
        _business_validation_errors(
            period,
            tuple(loans_with_rows),
            tuple(note_movements),
            set(),
        )
    )

    return ValidationOutcome(
        tuple(loan for _, loan in loans_with_rows),
        tuple(movement for _, movement in note_movements),
        sort_errors(errors),
    )
