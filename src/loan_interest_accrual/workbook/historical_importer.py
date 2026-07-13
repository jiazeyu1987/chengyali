from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl.workbook.workbook import Workbook

from loan_interest_accrual.domain import DayCountBasis, Loan, NaturalMonth

from .schema import WorkbookError, WorkbookErrorCode
from .validation import ValidationOutcome, sort_errors


_HISTORICAL_PROFILE_SHEETS = frozenset(
    {
        "24全年",
        "25年三季度子公司利息",
        "子公司借款2026年",
        "子公司借款利息（农商）",
        "子公司借款汇总22-24年",
        "子公司借款汇总25年",
        "银行借款360天算银行版本",
    }
)
_RATE_BASIS_FORMULA = re.compile(
    r"\*\s*([0-9]+(?:\.[0-9]+)?%?|0?\.[0-9]+)\s*/\s*(360|365)"
)
_RATE_BASIS_HEADER = re.compile(
    r"([0-9]+(?:\.[0-9]+)?)\s*%\D*(360|365)"
)
_FULL_YEAR = re.compile(r"(20\d{2})年")
_SHORT_YEAR = re.compile(r"(?<!\d)(2[0-9])(?:年|全年|年)")
_MONTH = re.compile(r"([1-9]|1[0-2])月")


@dataclass(frozen=True, slots=True)
class _HistoricalBlock:
    sheet_name: str
    title_row: int | None
    header_row: int
    month: int
    year: int


def is_historical_workbook(workbook: Workbook) -> bool:
    return any(
        sheet_name in workbook.sheetnames for sheet_name in _HISTORICAL_PROFILE_SHEETS
    )


def validate_historical_workbook(
    workbook: Workbook,
    period: NaturalMonth,
) -> ValidationOutcome:
    blocks = tuple(_matching_blocks(workbook, period))
    if not blocks:
        return ValidationOutcome(
            (),
            (),
            (
                WorkbookError(
                    WorkbookErrorCode.HISTORICAL_PERIOD_NOT_FOUND,
                    "历史工作簿",
                    None,
                    "计算月份",
                    f"历史工作簿中未找到 {period.year:04d}-{period.month:02d} 的可计算数据",
                ),
            ),
        )

    errors: list[WorkbookError] = []
    loans: list[Loan] = []
    for block in blocks:
        block_loans, block_errors = _parse_block(workbook, block, period)
        loans.extend(block_loans)
        errors.extend(block_errors)

    if errors:
        return ValidationOutcome((), (), sort_errors(errors))
    if not loans:
        return ValidationOutcome(
            (),
            (),
            (
                WorkbookError(
                    WorkbookErrorCode.HISTORICAL_PERIOD_NOT_FOUND,
                    "历史工作簿",
                    None,
                    "计算月份",
                    f"历史工作簿中未找到 {period.year:04d}-{period.month:02d} 的可计算数据行",
                ),
            ),
        )
    return ValidationOutcome(tuple(loans), (), ())


def _matching_blocks(workbook: Workbook, period: NaturalMonth):
    for sheet_name in sorted(_HISTORICAL_PROFILE_SHEETS):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in range(1, sheet.max_row + 1):
            if not _is_historical_header(sheet, row):
                continue
            month = _month_from_header(sheet.cell(row, 6).value)
            if month is None:
                continue
            title_row = row - 1 if row > 1 else None
            title_value = sheet.cell(title_row, 1).value if title_row is not None else None
            title_month = _month_from_title(title_value)
            if title_month is not None:
                month = title_month
            year = _year_from_context(sheet, title_row, row, month)
            if year == period.year and month == period.month:
                yield _HistoricalBlock(
                    sheet_name=sheet_name,
                    title_row=title_row,
                    header_row=row,
                    month=month,
                    year=year,
                )


def _is_historical_header(sheet, row: int) -> bool:
    values = [sheet.cell(row, column).value for column in range(1, 8)]
    return (
        values[0] == "序号"
        and type(values[1]) is str
        and "借款" in values[1]
        and values[2] == "借款时间"
        and values[3] == "借款金额"
        and type(values[5]) is str
        and "应计提天数" in values[5]
        and type(values[6]) is str
        and "计提利息" in values[6]
    )


def _month_from_header(value: object) -> int | None:
    if type(value) is not str:
        return None
    found = _MONTH.search(value)
    return int(found.group(1)) if found else None


def _month_from_title(value: object) -> int | None:
    if type(value) is not str or "借款" not in value:
        return None
    found = _MONTH.search(value)
    return int(found.group(1)) if found else None


def _year_from_context(sheet, title_row: int | None, header_row: int, month: int) -> int:
    title = sheet.cell(title_row, 1).value if title_row is not None else None
    for value in (title, sheet.title):
        year = _year_from_text(value)
        if year is not None:
            return year
    inferred = _infer_year_from_block_dates(sheet, header_row, month)
    if inferred is not None:
        return inferred
    return 0


def _year_from_text(value: object) -> int | None:
    if type(value) is not str:
        return None
    full = _FULL_YEAR.search(value)
    if full:
        return int(full.group(1))
    short = _SHORT_YEAR.search(value)
    if short:
        return 2000 + int(short.group(1))
    return None


def _infer_year_from_block_dates(sheet, header_row: int, month: int) -> int | None:
    for row in range(header_row + 1, min(sheet.max_row, header_row + 60) + 1):
        if _is_historical_header(sheet, row):
            break
        value = _date_value(sheet.cell(row, 3).value)
        if value is not None and value.month <= month:
            return value.year
    return None


def _parse_block(
    workbook: Workbook,
    block: _HistoricalBlock,
    period: NaturalMonth,
) -> tuple[list[Loan], list[WorkbookError]]:
    sheet = workbook[block.sheet_name]
    loans: list[Loan] = []
    errors: list[WorkbookError] = []
    header_rate_basis = _rate_basis_from_header(sheet.cell(block.header_row, 5).value)

    for row in range(block.header_row + 1, sheet.max_row + 1):
        if _is_historical_header(sheet, row):
            break
        if _starts_next_title(sheet, row):
            break
        if _ends_block(sheet, row):
            break
        if _is_blank_data_row(sheet, row):
            continue
        sequence = sheet.cell(row, 1).value
        if type(sequence) is not int:
            continue

        company_or_bank = _text(sheet.cell(row, 2).value)
        borrowed_at = _date_value(sheet.cell(row, 3).value)
        principal = _decimal(sheet.cell(row, 4).value)
        rate_basis = _rate_basis_from_formula(
            sheet.cell(row, 5).value
        ) or header_rate_basis
        if company_or_bank is None:
            errors.append(_error(block.sheet_name, row, "借款", "借款不能为空"))
            continue
        if borrowed_at is None:
            errors.append(_error(block.sheet_name, row, "借款时间", "借款时间必须是有效日期"))
            continue
        if borrowed_at > period.end_date:
            continue
        if principal is None or principal <= Decimal("0"):
            errors.append(_error(block.sheet_name, row, "借款金额", "借款金额必须大于0"))
            continue
        if rate_basis is None:
            errors.append(_error(block.sheet_name, row, "年利率", "无法识别历史表中的年利率和计息基准"))
            continue
        annual_rate, basis = rate_basis
        loans.append(
            Loan(
                loan_id=f"历史:{block.sheet_name}:{row}",
                company_name=_company_name(block.sheet_name, company_or_bank),
                contract_number=f"历史文件-{block.sheet_name}-第{row}行",
                bank_name=_bank_name(block.sheet_name, company_or_bank),
                opening_principal=principal,
                annual_rate=annual_rate,
                day_count_basis=basis,
                accrual_start=borrowed_at,
                accrual_end=period.end_date,
                capitalize_interest=False,
            )
        )
    return loans, errors


def _error(sheet: str, row: int, field: str, message: str) -> WorkbookError:
    return WorkbookError(
        WorkbookErrorCode.VALUE_TYPE_INVALID,
        sheet,
        row,
        field,
        message,
    )


def _text(value: object) -> str | None:
    return value.strip() if type(value) is str and value.strip() else None


def _decimal(value: object) -> Decimal | None:
    if type(value) is bool or not isinstance(value, (int, float, Decimal)):
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return result if result.is_finite() else None


def _date_value(value: object) -> date | None:
    if type(value) is datetime:
        return value.date()
    if type(value) is date:
        return value
    if type(value) is str:
        cleaned = value.strip()
        for pattern in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(cleaned, pattern).date()
            except ValueError:
                pass
    return None


def _rate_basis_from_formula(value: object) -> tuple[Decimal, DayCountBasis] | None:
    if type(value) is not str:
        return None
    found = _RATE_BASIS_FORMULA.search(value.replace(" ", ""))
    if not found:
        return None
    return _rate_basis(found.group(1), found.group(2))


def _rate_basis_from_header(value: object) -> tuple[Decimal, DayCountBasis] | None:
    if type(value) is not str:
        return None
    found = _RATE_BASIS_HEADER.search(value)
    if not found:
        return None
    return _rate_basis(found.group(1) + "%", found.group(2))


def _rate_basis(rate_text: str, basis_text: str) -> tuple[Decimal, DayCountBasis] | None:
    text = rate_text.strip()
    if text.endswith("%"):
        annual_rate = Decimal(text[:-1]) / Decimal("100")
    else:
        annual_rate = Decimal(text)
    if annual_rate <= Decimal("0") or annual_rate >= Decimal("1"):
        return None
    return annual_rate, DayCountBasis(int(basis_text))


def _starts_next_title(sheet, row: int) -> bool:
    value = sheet.cell(row, 1).value
    return type(value) is str and "月" in value and "借款" in value


def _ends_block(sheet, row: int) -> bool:
    values = [sheet.cell(row, column).value for column in range(1, 5)]
    return any(
        type(value) is str and ("合计" in value or "小计" in value or "制单" in value)
        for value in values
    )


def _is_blank_data_row(sheet, row: int) -> bool:
    return all(sheet.cell(row, column).value in (None, "") for column in range(1, 8))


def _company_name(sheet_name: str, label: str) -> str:
    if "银行借款" in sheet_name:
        return "银行借款"
    return label


def _bank_name(sheet_name: str, label: str) -> str:
    if "银行借款" in sheet_name:
        return label
    return "历史工作簿"
