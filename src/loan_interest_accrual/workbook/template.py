from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .limits import MAX_LOAN_ROWS, MAX_MOVEMENT_ROWS
from .schema import (
    CURRENCY_NUMBER_FORMAT,
    DATE_NUMBER_FORMAT,
    LOAN_SHEET,
    LOAN_TEMPLATE_HEADERS,
    LOAN_REQUIRED_HEADERS,
    PERCENT_NUMBER_FORMAT,
)


_LOAN_INSTRUCTIONS = (
    "必填公司名称。",
    "选填贷款合同号。",
    "选填贷款银行；填写后将在预览和导出结果中显示。",
    "必填人民币元金额，必须大于或等于0。",
    "必填Excel百分比，例如2.50%；区间应提利息按360天基准计算。",
    "必填有效日期；借款当天不计息，从次日开始计息。",
    "可选备注；“26.3.27还本金”表示当日全额还清；“26.6.9还700万”表示当日仍按原本金计息，次日起本金减少700万元。",
)


def _configure_sheet(sheet, headers, instructions, max_rows: int) -> None:
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).column_letter}1"
    for column, (header, instruction) in enumerate(
        zip(headers, instructions, strict=True), start=1
    ):
        cell = sheet.cell(1, column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="B7791F" if header in LOAN_REQUIRED_HEADERS else "374151",
        )
        cell.comment = Comment(f"填写说明：{instruction}", "本地应用")
        sheet.column_dimensions[cell.column_letter].width = max(14, len(header) * 2 + 2)
    for row in range(2, 3):
        for column in range(1, len(headers) + 1):
            sheet.cell(row, column)
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).column_letter}{max_rows + 1}"


def generate_standard_template() -> bytes:
    workbook = Workbook()
    loan_sheet = workbook.active
    loan_sheet.title = LOAN_SHEET
    _configure_sheet(
        loan_sheet, LOAN_TEMPLATE_HEADERS, _LOAN_INSTRUCTIONS, MAX_LOAN_ROWS
    )

    positions = {
        header: get_column_letter(index)
        for index, header in enumerate(LOAN_TEMPLATE_HEADERS, start=1)
    }
    loan_sheet[f"{positions['期初本金（元）']}2"].number_format = CURRENCY_NUMBER_FORMAT
    loan_sheet[f"{positions['年利率']}2"].number_format = PERCENT_NUMBER_FORMAT
    loan_sheet[f"{positions['借款时间']}2"].number_format = DATE_NUMBER_FORMAT

    principal = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
    )
    rate_column = positions["年利率"]
    rate = DataValidation(
        type="custom",
        formula1=f"=AND({rate_column}2>0,{rate_column}2<1)",
        allow_blank=False,
    )
    start_date = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="DATE(1900,1,1)",
        allow_blank=False,
    )
    for validation in (principal, rate, start_date):
        loan_sheet.add_data_validation(validation)
    principal.add(f"{positions['期初本金（元）']}2:{positions['期初本金（元）']}{MAX_LOAN_ROWS + 1}")
    rate.add(f"{rate_column}2:{rate_column}{MAX_LOAN_ROWS + 1}")
    start_date.add(f"{positions['借款时间']}2:{positions['借款时间']}{MAX_LOAN_ROWS + 1}")

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
