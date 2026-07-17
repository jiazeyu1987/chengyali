from __future__ import annotations

from collections import defaultdict
from collections.abc import Iterable
from datetime import datetime, timezone
from io import BytesIO
import locale
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .amortization_schema import (
    AMORTIZATION_RESULT_HEADERS,
    AMOUNT_NUMBER_FORMAT,
    DATE_NUMBER_FORMAT,
    MONTH_COUNT_NUMBER_FORMAT,
)

if TYPE_CHECKING:
    from loan_interest_accrual.application.amortization_results import (
        AmortizationCalculationResult,
    )


RESULT_SHEET = "摊销明细"
CLASSIFIED_SHEET = "分类明细"
RESULT_SHEETS = (
    RESULT_SHEET,
    CLASSIFIED_SHEET,
)

VISIBLE_COLUMN_COUNT = len(AMORTIZATION_RESULT_HEADERS)
FIRST_DATA_ROW = 3
MONTH_NUMBER_FORMAT = 'yy.m"月"'
SUMMARY_AMOUNT_NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00);0'

try:
    locale.setlocale(locale.LC_COLLATE, "Chinese_China.936")
except locale.Error:
    pass


class AmortizationExportInvariantError(ValueError):
    pass


def _set_value(cell, value: object) -> None:
    cell.value = value
    if type(value) is str and not value.startswith("="):
        cell.data_type = "s"


def _collation_key(value: str) -> str:
    return locale.strxfrm(value.casefold())


def _classification_groups(
    rows: Iterable["AmortizationPreviewRow"],
) -> tuple[tuple[str, tuple[str, ...]], ...]:
    grouped: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        grouped[row.expense_category].add(row.primary_category)
    return tuple(
        (
            expense_category,
            tuple(sorted(grouped[expense_category], key=_collation_key)),
        )
        for expense_category in sorted(grouped, key=_collation_key)
    )


def _add_classification_summary_sheet(
    workbook: Workbook,
    calculation: "AmortizationCalculationResult",
    *,
    generated_at: datetime,
):
    sheet = workbook.create_sheet(CLASSIFIED_SHEET)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"

    period = calculation.period
    sheet.merge_cells("A1:C1")
    _set_value(
        sheet["A1"],
        f"无形资产、长期待摊费用摊销分类汇总  {period.year}年{period.month}月",
    )
    sheet["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor="C6E0B4")
    sheet.row_dimensions[1].height = 28

    sheet.cell(1, 5, "计提月份")
    sheet.cell(1, 6, period.start_date)
    sheet.cell(1, 6).number_format = DATE_NUMBER_FORMAT
    sheet.cell(2, 5, "生成时间")
    sheet.cell(2, 6, generated_at.replace(tzinfo=None))
    sheet.cell(2, 6).number_format = "yyyy-mm-dd hh:mm:ss"
    sheet.column_dimensions["E"].hidden = True
    sheet.column_dimensions["F"].hidden = True

    headers = ("费用", "一级分类", "当期应摊销金额")
    border = Border(
        left=Side(style="thin", color="A9D18E"),
        right=Side(style="thin", color="A9D18E"),
        top=Side(style="thin", color="A9D18E"),
        bottom=Side(style="thin", color="A9D18E"),
    )
    header_fill = PatternFill("solid", fgColor="C6E0B4")
    body_fill = PatternFill("solid", fgColor="E2F0D9")
    group_fill = PatternFill("solid", fgColor="D9EAD3")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[3].height = 24

    detail_first_row = FIRST_DATA_ROW
    detail_last_row = detail_first_row + len(calculation.rows) - 1
    current_row = 4
    for expense_category, primary_categories in _classification_groups(
        calculation.rows
    ):
        expense_row = current_row
        _set_value(sheet.cell(current_row, 1), expense_category)
        sheet.cell(
            current_row,
            3,
            f"=SUMIF('{RESULT_SHEET}'!$E${detail_first_row}:$E${detail_last_row},A{current_row},'{RESULT_SHEET}'!$N${detail_first_row}:$N${detail_last_row})",
        )
        for column in range(1, 4):
            cell = sheet.cell(current_row, column)
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
            cell.fill = group_fill
            cell.border = border
        current_row += 1
        for primary_category in primary_categories:
            _set_value(sheet.cell(current_row, 2), primary_category)
            sheet.cell(
                current_row,
                3,
                f"=SUMIFS('{RESULT_SHEET}'!$N${detail_first_row}:$N${detail_last_row},'{RESULT_SHEET}'!$E${detail_first_row}:$E${detail_last_row},$A${expense_row},'{RESULT_SHEET}'!$C${detail_first_row}:$C${detail_last_row},B{current_row})",
            )
            for column in range(1, 4):
                cell = sheet.cell(current_row, column)
                cell.font = Font(name="Microsoft YaHei", size=10)
                cell.fill = body_fill
                cell.border = border
            sheet.cell(current_row, 2).alignment = Alignment(
                horizontal="left",
                vertical="center",
                indent=1,
            )
            current_row += 1

    total_row = current_row
    _set_value(sheet.cell(total_row, 1), "总计")
    sheet.cell(
        total_row,
        3,
        f"=SUM('{RESULT_SHEET}'!$N${detail_first_row}:$N${detail_last_row})",
    )
    for column in range(1, 4):
        cell = sheet.cell(total_row, column)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.fill = header_fill
        cell.border = border

    for row_number in range(4, total_row + 1):
        sheet.cell(row_number, 1).alignment = Alignment(
            horizontal="left", vertical="center"
        )
        sheet.cell(row_number, 3).alignment = Alignment(
            horizontal="right", vertical="center"
        )
        sheet.cell(row_number, 3).number_format = SUMMARY_AMOUNT_NUMBER_FORMAT
        sheet.row_dimensions[row_number].height = 21

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 40
    sheet.column_dimensions["C"].width = 24
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:3"
    sheet.print_area = f"A1:C{total_row}"
    return sheet


def _add_result_sheet(
    workbook: Workbook,
    calculation: "AmortizationCalculationResult",
    *,
    sheet_name: str,
    title_suffix: str,
    rows: tuple["AmortizationPreviewRow", ...],
    generated_at: datetime,
):
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C3"

    period = calculation.period
    title = (
        f"无形资产、长期待摊费用摊销明细{title_suffix}  "
        f"{period.year}年{period.month}月"
    )
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=VISIBLE_COLUMN_COUNT)
    _set_value(sheet["A1"], title)
    sheet["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor="C6E0B4")
    sheet.row_dimensions[1].height = 28

    input_fill = PatternFill("solid", fgColor="FFF200")
    formula_fill = PatternFill("solid", fgColor="C6E0B4")
    formula_data_fill = PatternFill("solid", fgColor="E2F0D9")
    border = Border(
        left=Side(style="thin", color="548235"),
        right=Side(style="thin", color="548235"),
        top=Side(style="thin", color="548235"),
        bottom=Side(style="thin", color="548235"),
    )
    for column, header in enumerate(AMORTIZATION_RESULT_HEADERS, start=1):
        cell = sheet.cell(2, column, header)
        cell.fill = formula_fill if column == 1 or column >= 11 else input_fill
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[2].height = 34

    metadata_label_column = VISIBLE_COLUMN_COUNT + 1
    metadata_value_column = VISIBLE_COLUMN_COUNT + 2
    metadata_label_letter = get_column_letter(metadata_label_column)
    metadata_value_letter = get_column_letter(metadata_value_column)
    sheet.cell(1, metadata_label_column, "计提月份")
    sheet.cell(1, metadata_value_column, period.start_date)
    sheet.cell(1, metadata_value_column).number_format = DATE_NUMBER_FORMAT
    sheet.cell(2, metadata_label_column, "生成时间")
    sheet.cell(2, metadata_value_column, generated_at.replace(tzinfo=None))
    sheet.cell(2, metadata_value_column).number_format = "yyyy-mm-dd hh:mm:ss"
    sheet.column_dimensions[metadata_label_letter].hidden = True
    sheet.column_dimensions[metadata_value_letter].hidden = True

    for offset, row in enumerate(rows):
        excel_row = FIRST_DATA_ROW + offset
        values = (
            period.start_date,
            row.sequence,
            row.primary_category,
            row.name,
            row.expense_category,
            row.original_value,
            row.residual_value,
            row.amortization_start,
            row.booking_month,
            row.amortization_term_months,
        )
        for column, value in enumerate(values, start=1):
            _set_value(sheet.cell(excel_row, column), value)
        sheet.cell(excel_row, 11, f"=ROUND((F{excel_row}-G{excel_row})/J{excel_row},2)")
        sheet.cell(
            excel_row,
            12,
            f"=MIN(J{excel_row},MAX(0,(YEAR(${metadata_value_letter}$1)-YEAR(H{excel_row}))*12+MONTH(${metadata_value_letter}$1)-MONTH(H{excel_row})+1))",
        )
        sheet.cell(
            excel_row,
            13,
            f"=IF(L{excel_row}>=J{excel_row},F{excel_row},K{excel_row}*L{excel_row})",
        )
        sheet.cell(excel_row, 14, f'=IF(L{excel_row}>=J{excel_row},"",K{excel_row})')
        sheet.cell(excel_row, 15, f'=IF(L{excel_row}>=J{excel_row},"",K{excel_row})')
        sheet.cell(excel_row, 16, None)
        sheet.cell(excel_row, 17, f'=IF(L{excel_row}>=J{excel_row},"",F{excel_row}-M{excel_row})')
        for column in range(1, VISIBLE_COLUMN_COUNT + 1):
            cell = sheet.cell(excel_row, column)
            cell.border = border
            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column in (3, 4, 5)
                    else "center"
                    if column == 1
                    else "right"
                ),
                vertical="center",
            )
            if column == 1 or column >= 11:
                cell.fill = formula_data_fill

    last_data_row = FIRST_DATA_ROW + len(rows) - 1
    total_row = last_data_row + 1
    sheet.cell(total_row, 3, "合计")
    for column in (6, 11, 13, 14, 15, 17):
        letter = get_column_letter(column)
        sheet.cell(total_row, column, f"=SUM({letter}{FIRST_DATA_ROW}:{letter}{last_data_row})")
    for column in (8, 9, 10, 12):
        sheet.cell(total_row, column, "/")
    for column in range(1, VISIBLE_COLUMN_COUNT + 1):
        cell = sheet.cell(total_row, column)
        cell.fill = PatternFill("solid", fgColor="C6E0B4")
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if column in (1, 2, 3, 8, 9, 10, 12, 16) else "right",
            vertical="center",
        )

    for excel_row in range(FIRST_DATA_ROW, total_row + 1):
        sheet.cell(excel_row, 1).number_format = MONTH_NUMBER_FORMAT
        for column in (6, 7, 11, 13, 14, 15, 16, 17):
            sheet.cell(excel_row, column).number_format = AMOUNT_NUMBER_FORMAT
        for column in (8, 9):
            sheet.cell(excel_row, column).number_format = DATE_NUMBER_FORMAT
        for column in (2, 10, 12):
            sheet.cell(excel_row, column).number_format = MONTH_COUNT_NUMBER_FORMAT

    widths = (11, 8, 20, 42, 16, 16, 13, 17, 17, 17, 15, 22, 23, 19, 21, 12, 16)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = f"A2:Q{last_data_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:2"
    sheet.print_area = f"A1:Q{total_row}"
    return sheet


def export_amortization_workbook(
    calculation: "AmortizationCalculationResult",
    *,
    generated_at: datetime | None = None,
) -> bytes:
    if any(not check.passed for check in calculation.checks):
        raise AmortizationExportInvariantError(
            "amortization checks failed; export was blocked"
        )
    if not calculation.rows:
        raise AmortizationExportInvariantError("no amortization rows to export")

    generated_at = generated_at or datetime.now(timezone.utc)
    workbook = Workbook()
    workbook.remove(workbook.active)
    detail_sheet = _add_result_sheet(
        workbook,
        calculation,
        sheet_name=RESULT_SHEET,
        title_suffix="",
        rows=calculation.rows,
        generated_at=generated_at,
    )
    classified_sheet = _add_classification_summary_sheet(
        workbook,
        calculation,
        generated_at=generated_at,
    )
    workbook.active = 0
    detail_sheet.sheet_properties.tabColor = "548235"
    classified_sheet.sheet_properties.tabColor = "70AD47"

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()
